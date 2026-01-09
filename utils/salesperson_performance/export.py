# utils/salesperson_performance/export.py
"""
Formatted Excel Export for Salesperson Performance

Creates professional Excel reports with:
- Cover sheet with KPI summary
- Pipeline & Forecast metrics
- KPI Breakdown by Type (NEW v2.3.0)
- Salesperson summary with conditional formatting
- Monthly breakdown
- Detailed transactions
- Backlog summary, detail, and by ETD (multi-year support)

Uses openpyxl for formatting capabilities.

CHANGELOG:
- v2.3.0: SYNCED Excel export with UI logic
          - Added Overall Achievement to Summary sheet
          - NEW: KPI Breakdown sheet showing each KPI type contribution
          - NEW: Complex KPIs with targets and achievements
          - NEW: Multi-year Backlog by ETD support
          - Summary sheet now matches Overview tab exactly
- v2.2.0: UPDATED By Salesperson sheet to use overall_achievement
          - Prefer overall_achievement over revenue_achievement
          - overall_achievement = weighted avg of all KPIs (same as Overview)
          - Conditional formatting now works with overall_achievement
- v2.0.0: ADDED Comprehensive Export with full data
          - New create_comprehensive_report() method
          - Pipeline & Forecast sheet with Revenue/GP/GP1 metrics
          - Backlog Summary by salesperson
          - Backlog Detail with all line items
          - Backlog by ETD Month
          - Export button moved to main page level
- v1.0.0: Initial version with basic export

VERSION: 2.3.0
"""

import logging
from datetime import datetime
from io import BytesIO
from typing import Dict, List, Optional
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill, NamedStyle
)
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule

from .constants import EXCEL_STYLES, MONTH_ORDER

logger = logging.getLogger(__name__)


class SalespersonExport:
    """
    Excel report generator for salesperson performance.
    
    Usage:
        exporter = SalespersonExport()
        excel_bytes = exporter.create_report(
            summary_df=summary_df,
            monthly_df=monthly_df,
            metrics=metrics,
            filters=filters
        )
        
        st.download_button(
            label="Download Report",
            data=excel_bytes,
            file_name="salesperson_performance.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    """
    
    def __init__(self):
        """Initialize with default styles."""
        self.wb = None
        self._init_styles()
    
    def _init_styles(self):
        """Initialize reusable styles."""
        # Colors
        self.header_fill = PatternFill(
            start_color=EXCEL_STYLES['header_fill_color'],
            end_color=EXCEL_STYLES['header_fill_color'],
            fill_type='solid'
        )
        
        self.header_font = Font(
            bold=True,
            color=EXCEL_STYLES['header_font_color'],
            size=11
        )
        
        self.title_font = Font(bold=True, size=16)
        self.subtitle_font = Font(bold=True, size=12)
        self.normal_font = Font(size=11)
        
        # Borders
        thin_border = Side(style='thin', color='000000')
        self.cell_border = Border(
            left=thin_border,
            right=thin_border,
            top=thin_border,
            bottom=thin_border
        )
        
        # Alignments
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.right_align = Alignment(horizontal='right', vertical='center')
        self.left_align = Alignment(horizontal='left', vertical='center')
        
        # Number formats
        self.currency_format = EXCEL_STYLES['currency_format']
        self.percent_format = EXCEL_STYLES['percent_format']
        
        # Achievement colors
        self.achievement_good_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        self.achievement_warning_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        self.achievement_bad_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    def _get_achievement_fill(self, achievement: float) -> PatternFill:
        """Get fill color based on achievement percentage."""
        if achievement is None:
            return None
        if achievement >= 100:
            return self.achievement_good_fill
        elif achievement >= 80:
            return self.achievement_warning_fill
        else:
            return self.achievement_bad_fill
    
    # =========================================================================
    # MAIN EXPORT METHOD
    # =========================================================================
    
    def create_report(
        self,
        summary_df: pd.DataFrame,
        monthly_df: pd.DataFrame,
        metrics: Dict,
        filters: Dict,
        detail_df: pd.DataFrame = None,
        complex_kpis: Dict = None,
        yoy_metrics: Dict = None
    ) -> BytesIO:
        """
        Create formatted Excel report with multiple sheets.
        
        Args:
            summary_df: Salesperson summary data
            monthly_df: Monthly breakdown data
            metrics: Overview metrics dict
            filters: Filter settings dict
            detail_df: Optional detailed transactions
            complex_kpis: Optional complex KPI metrics
            yoy_metrics: Optional YoY comparison metrics
            
        Returns:
            BytesIO containing Excel file
        """
        self.wb = Workbook()
        
        # Create sheets
        self._create_cover_sheet(metrics, filters, complex_kpis, yoy_metrics)
        self._create_summary_sheet(summary_df)
        self._create_monthly_sheet(monthly_df)
        
        if detail_df is not None and not detail_df.empty:
            self._create_detail_sheet(detail_df)
        
        # Remove default empty sheet if exists
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']
        
        # Save to BytesIO
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        
        logger.info("Excel report created successfully")
        return output
    
    # =========================================================================
    # COMPREHENSIVE REPORT (UPDATED v2.3.0)
    # =========================================================================
    
    def create_comprehensive_report(
        self,
        # Summary metrics
        metrics: Dict,
        complex_kpis: Dict,
        pipeline_metrics: Dict,
        filters: Dict,
        yoy_metrics: Dict = None,
        
        # NEW v2.3.0: Overall Achievement data (synced with UI)
        overall_achievement_data: Dict = None,
        
        # Salesperson & Monthly data
        salesperson_summary_df: pd.DataFrame = None,
        monthly_df: pd.DataFrame = None,
        
        # Sales Detail
        sales_detail_df: pd.DataFrame = None,
        
        # Backlog data
        backlog_summary_df: pd.DataFrame = None,
        backlog_detail_df: pd.DataFrame = None,
        backlog_by_month_df: pd.DataFrame = None,
        in_period_backlog_analysis: Dict = None,
        
        # NEW v2.3.0: Multi-year backlog support
        backlog_by_month_multiyear_df: pd.DataFrame = None,
    ) -> BytesIO:
        """
        Create comprehensive Excel report with all performance data.
        
        UPDATED v2.3.0: Full sync with UI logic.
        
        Sheets:
        1. Summary - KPIs, filters, period info (with Overall Achievement)
        2. KPI Breakdown - KPI type contributions to Overall Achievement (NEW)
        3. Pipeline & Forecast - Revenue/GP/GP1 forecast vs target
        4. By Salesperson - Summary by salesperson
        5. Monthly Trend - Monthly breakdown
        6. Sales Detail - Transaction list
        7. Backlog Summary - Backlog by salesperson
        8. Backlog Detail - Backlog line items
        9. Backlog by ETD - Backlog grouped by ETD month (multi-year)
        
        Args:
            metrics: Overview metrics from calculate_overview_metrics()
            complex_kpis: Complex KPIs (new customers, products, business)
            pipeline_metrics: Pipeline & Forecast from calculate_pipeline_forecast_metrics()
            filters: Current filter settings
            yoy_metrics: Optional YoY comparison
            overall_achievement_data: From calculate_overall_kpi_achievement() (NEW)
            salesperson_summary_df: Summary by salesperson
            monthly_df: Monthly breakdown
            sales_detail_df: Sales transactions
            backlog_summary_df: Backlog aggregated by salesperson
            backlog_detail_df: Backlog line items
            backlog_by_month_df: Backlog by ETD month (single year, legacy)
            in_period_backlog_analysis: In-period analysis with overdue info
            backlog_by_month_multiyear_df: Backlog by ETD month (multi-year, NEW)
            
        Returns:
            BytesIO containing Excel file
        """
        self.wb = Workbook()
        
        # Sheet 1: Summary (Cover Page) - UPDATED with Overall Achievement
        self._create_comprehensive_cover_sheet(
            metrics, complex_kpis, pipeline_metrics, filters, yoy_metrics,
            overall_achievement_data  # NEW parameter
        )
        
        # Sheet 2: KPI Breakdown (NEW v2.3.0)
        if overall_achievement_data and overall_achievement_data.get('kpi_details'):
            self._create_kpi_breakdown_sheet(overall_achievement_data)
        
        # Sheet 3: Pipeline & Forecast
        if pipeline_metrics:
            self._create_pipeline_forecast_sheet(pipeline_metrics)
        
        # Sheet 4: By Salesperson
        if salesperson_summary_df is not None and not salesperson_summary_df.empty:
            self._create_summary_sheet(salesperson_summary_df)
        
        # Sheet 5: Monthly Trend
        if monthly_df is not None and not monthly_df.empty:
            self._create_monthly_sheet(monthly_df)
        
        # Sheet 6: Sales Detail
        if sales_detail_df is not None and not sales_detail_df.empty:
            self._create_sales_detail_sheet(sales_detail_df)
        
        # Sheet 7: Backlog Summary
        if backlog_summary_df is not None and not backlog_summary_df.empty:
            self._create_backlog_summary_sheet(backlog_summary_df)
        
        # Sheet 8: Backlog Detail
        if backlog_detail_df is not None and not backlog_detail_df.empty:
            self._create_backlog_detail_sheet(backlog_detail_df, in_period_backlog_analysis)
        
        # Sheet 9: Backlog by ETD - NEW v2.3.0: Prefer multi-year if available
        if backlog_by_month_multiyear_df is not None and not backlog_by_month_multiyear_df.empty:
            self._create_backlog_by_month_sheet_multiyear(backlog_by_month_multiyear_df)
        elif backlog_by_month_df is not None and not backlog_by_month_df.empty:
            self._create_backlog_by_month_sheet(backlog_by_month_df)
        
        # Remove default empty sheet if exists
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']
        
        # Save to BytesIO
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        
        logger.info("Comprehensive Excel report created successfully")
        return output
    
    # =========================================================================
    # COMPREHENSIVE COVER SHEET (UPDATED v2.3.0)
    # =========================================================================
    
    def _create_comprehensive_cover_sheet(
        self,
        metrics: Dict,
        complex_kpis: Dict,
        pipeline_metrics: Dict,
        filters: Dict,
        yoy_metrics: Dict = None,
        overall_achievement_data: Dict = None  # NEW v2.3.0
    ):
        """
        Create comprehensive cover page with all KPI summaries.
        
        UPDATED v2.3.0: Added Overall Achievement and synced with UI.
        """
        ws = self.wb.active
        ws.title = "Summary"
        
        row = 1
        
        # Title
        ws.cell(row=row, column=1, value="Salesperson Performance Report")
        ws.cell(row=row, column=1).font = self.title_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 2
        
        # Report Info
        ws.cell(row=row, column=1, value="Report Period:")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"{filters.get('period_type', 'YTD')} {filters.get('year', '')}")
        row += 1
        
        ws.cell(row=row, column=1, value="Date Range:")
        ws.cell(row=row, column=1).font = Font(bold=True)
        start_date = filters.get('start_date', '')
        end_date = filters.get('end_date', '')
        ws.cell(row=row, column=2, value=f"{start_date} to {end_date}")
        row += 1
        
        # Salespeople filter
        employee_ids = filters.get('employee_ids', [])
        ws.cell(row=row, column=1, value="Salespeople:")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"{len(employee_ids)} selected" if employee_ids else "All")
        row += 1
        
        ws.cell(row=row, column=1, value="Generated:")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M'))
        row += 2
        
        # ===== OVERALL ACHIEVEMENT SECTION (NEW v2.3.0) =====
        if overall_achievement_data and overall_achievement_data.get('overall_achievement') is not None:
            ws.cell(row=row, column=1, value="OVERALL ACHIEVEMENT")
            ws.cell(row=row, column=1).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            ws.cell(row=row, column=1).font = Font(bold=True, color='FFFFFF', size=12)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            row += 1
            
            overall_ach = overall_achievement_data['overall_achievement']
            total_weight = overall_achievement_data.get('total_weight', 0)
            kpi_count = overall_achievement_data.get('kpi_count', 0)
            is_single = overall_achievement_data.get('is_single_person', False)
            
            # Overall Achievement value with color coding
            ws.cell(row=row, column=1, value="Overall Achievement:")
            ws.cell(row=row, column=1).font = Font(bold=True, size=14)
            cell = ws.cell(row=row, column=2, value=f"{overall_ach:.1f}%")
            cell.font = Font(bold=True, size=14)
            cell.alignment = self.right_align
            cell.fill = self._get_achievement_fill(overall_ach)
            row += 1
            
            ws.cell(row=row, column=1, value="KPI Types Included:")
            ws.cell(row=row, column=2, value=f"{kpi_count}")
            ws.cell(row=row, column=2).alignment = self.right_align
            row += 1
            
            ws.cell(row=row, column=1, value="Total Weight:")
            ws.cell(row=row, column=2, value=f"{total_weight}")
            ws.cell(row=row, column=2).alignment = self.right_align
            row += 1
            
            weight_type = "Assignment Weight" if is_single else "KPI Type Default Weight"
            ws.cell(row=row, column=1, value="Weight Source:")
            ws.cell(row=row, column=2, value=weight_type)
            ws.cell(row=row, column=2).alignment = self.right_align
            row += 2
        
        # ===== SALES PERFORMANCE SECTION =====
        ws.cell(row=row, column=1, value="SALES PERFORMANCE")
        ws.cell(row=row, column=1).font = self.subtitle_font
        ws.cell(row=row, column=1).fill = self.header_fill
        ws.cell(row=row, column=1).font = Font(bold=True, color='FFFFFF')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1
        
        kpi_rows = [
            ("Total Revenue", f"${metrics.get('total_revenue', 0):,.0f}"),
            ("Gross Profit", f"${metrics.get('total_gp', 0):,.0f}"),
            ("GP1", f"${metrics.get('total_gp1', 0):,.0f}"),
            ("GP Margin", f"{metrics.get('gp_percent', 0):.1f}%"),
            ("Total Customers", f"{metrics.get('total_customers', 0):,}"),
            ("Total Invoices", f"{metrics.get('total_invoices', 0):,}"),
            ("Total Orders", f"{metrics.get('total_orders', 0):,}"),
        ]
        
        # Add individual achievements
        if metrics.get('revenue_achievement'):
            kpi_rows.append(("Revenue Achievement", f"{metrics['revenue_achievement']:.1f}%"))
        if metrics.get('gp_achievement'):
            kpi_rows.append(("GP Achievement", f"{metrics['gp_achievement']:.1f}%"))
        if metrics.get('gp1_achievement'):
            kpi_rows.append(("GP1 Achievement", f"{metrics['gp1_achievement']:.1f}%"))
        
        for label, value in kpi_rows:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=2).alignment = self.right_align
            row += 1
        
        row += 1
        
        # ===== COMPLEX KPIs SECTION (UPDATED v2.3.0 - with targets & achievements) =====
        if complex_kpis:
            ws.cell(row=row, column=1, value="COMPLEX KPIs")
            ws.cell(row=row, column=1).fill = self.header_fill
            ws.cell(row=row, column=1).font = Font(bold=True, color='FFFFFF')
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            row += 1
            
            # New Customers
            nc_actual = complex_kpis.get('new_customer_count', 0)
            nc_target = complex_kpis.get('new_customer_target')
            nc_ach = complex_kpis.get('new_customer_achievement')
            
            ws.cell(row=row, column=1, value="New Customers")
            if nc_target:
                ws.cell(row=row, column=2, value=f"{nc_actual:.1f} / {nc_target:.1f} ({nc_ach:.1f}%)" if nc_ach else f"{nc_actual:.1f} / {nc_target:.1f}")
            else:
                ws.cell(row=row, column=2, value=f"{nc_actual:.1f}")
            ws.cell(row=row, column=2).alignment = self.right_align
            row += 1
            
            # New Products
            np_actual = complex_kpis.get('new_product_count', 0)
            np_target = complex_kpis.get('new_product_target')
            np_ach = complex_kpis.get('new_product_achievement')
            
            ws.cell(row=row, column=1, value="New Products")
            if np_target:
                ws.cell(row=row, column=2, value=f"{np_actual:.1f} / {np_target:.1f} ({np_ach:.1f}%)" if np_ach else f"{np_actual:.1f} / {np_target:.1f}")
            else:
                ws.cell(row=row, column=2, value=f"{np_actual:.1f}")
            ws.cell(row=row, column=2).alignment = self.right_align
            row += 1
            
            # New Combos (if available)
            if 'num_new_combos' in complex_kpis:
                combo_actual = complex_kpis.get('num_new_combos', 0)
                combo_target = complex_kpis.get('new_combo_target')
                combo_ach = complex_kpis.get('new_combo_achievement')
                
                ws.cell(row=row, column=1, value="New Combos")
                if combo_target:
                    ws.cell(row=row, column=2, value=f"{combo_actual:.1f} / {combo_target:.1f} ({combo_ach:.1f}%)" if combo_ach else f"{combo_actual:.1f} / {combo_target:.1f}")
                else:
                    ws.cell(row=row, column=2, value=f"{combo_actual:.1f}")
                ws.cell(row=row, column=2).alignment = self.right_align
                row += 1
            
            # New Business Revenue
            nb_actual = complex_kpis.get('new_business_revenue', 0)
            nb_target = complex_kpis.get('new_business_target')
            nb_ach = complex_kpis.get('new_business_achievement')
            
            ws.cell(row=row, column=1, value="New Business Revenue")
            if nb_target:
                ws.cell(row=row, column=2, value=f"${nb_actual:,.0f} / ${nb_target:,.0f} ({nb_ach:.1f}%)" if nb_ach else f"${nb_actual:,.0f} / ${nb_target:,.0f}")
            else:
                ws.cell(row=row, column=2, value=f"${nb_actual:,.0f}")
            ws.cell(row=row, column=2).alignment = self.right_align
            row += 1
            
            row += 1
        
        # ===== PIPELINE & FORECAST SUMMARY =====
        if pipeline_metrics:
            ws.cell(row=row, column=1, value="PIPELINE & FORECAST")
            ws.cell(row=row, column=1).fill = self.header_fill
            ws.cell(row=row, column=1).font = Font(bold=True, color='FFFFFF')
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            row += 1
            
            summary = pipeline_metrics.get('summary', {})
            revenue = pipeline_metrics.get('revenue', {})
            
            pipeline_rows = [
                ("Total Backlog (Revenue)", f"${summary.get('total_backlog_revenue', 0):,.0f}"),
                ("Total Backlog (GP)", f"${summary.get('total_backlog_gp', 0):,.0f}"),
                ("In-Period Backlog (Revenue KPI)", f"${revenue.get('in_period_backlog', 0):,.0f}"),
                ("Revenue Forecast", f"${revenue.get('forecast', 0):,.0f}" if revenue.get('forecast') else "N/A"),
                ("Revenue Target", f"${revenue.get('target', 0):,.0f}" if revenue.get('target') else "N/A"),
            ]
            
            gap = revenue.get('gap')
            if gap is not None:
                gap_label = "Surplus" if gap >= 0 else "GAP"
                pipeline_rows.append((f"Revenue {gap_label}", f"${gap:+,.0f}"))
            
            for label, value in pipeline_rows:
                ws.cell(row=row, column=1, value=label)
                ws.cell(row=row, column=2, value=value)
                ws.cell(row=row, column=2).alignment = self.right_align
                row += 1
            
            row += 1
        
        # ===== YoY COMPARISON =====
        if yoy_metrics:
            ws.cell(row=row, column=1, value="YEAR-OVER-YEAR COMPARISON")
            ws.cell(row=row, column=1).fill = self.header_fill
            ws.cell(row=row, column=1).font = Font(bold=True, color='FFFFFF')
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            row += 1
            
            yoy_rows = [
                ("Revenue YoY", yoy_metrics.get('total_revenue_yoy')),
                ("GP YoY", yoy_metrics.get('total_gp_yoy')),
                ("GP1 YoY", yoy_metrics.get('total_gp1_yoy')),
                ("Customers YoY", yoy_metrics.get('total_customers_yoy')),
            ]
            
            for label, value in yoy_rows:
                ws.cell(row=row, column=1, value=label)
                if value is not None:
                    ws.cell(row=row, column=2, value=f"{value:+.1f}%")
                else:
                    ws.cell(row=row, column=2, value="N/A")
                ws.cell(row=row, column=2).alignment = self.right_align
                row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
    
    # =========================================================================
    # KPI BREAKDOWN SHEET (NEW v2.3.0)
    # =========================================================================
    
    def _create_kpi_breakdown_sheet(self, overall_achievement_data: Dict):
        """
        Create KPI Breakdown sheet showing each KPI type's contribution.
        
        NEW v2.3.0: Syncs with UI KPI Progress display.
        
        Shows:
        - KPI Type name
        - Actual value
        - Target (Prorated)
        - Target (Annual)
        - Achievement %
        - Weight
        - Weighted Score
        - # Employees with this KPI
        """
        kpi_details = overall_achievement_data.get('kpi_details', [])
        
        if not kpi_details:
            return
        
        ws = self.wb.create_sheet("KPI Breakdown")
        
        # Title
        ws.cell(row=1, column=1, value="KPI Achievement Breakdown")
        ws.cell(row=1, column=1).font = self.title_font
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        
        # Subtitle with formula explanation
        is_single = overall_achievement_data.get('is_single_person', False)
        if is_single:
            formula_text = "Formula: Overall = Σ(Achievement × Assignment Weight) / Σ(Assignment Weight)"
        else:
            formula_text = "Formula: Overall = Σ(KPI Type Achievement × Default Weight) / Σ(Default Weight)"
        
        ws.cell(row=2, column=1, value=formula_text)
        ws.cell(row=2, column=1).font = Font(italic=True, color='666666')
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
        
        # Headers
        headers = [
            ('KPI Type', 20),
            ('Actual', 15),
            ('Target (Prorated)', 18),
            ('Target (Annual)', 15),
            ('Achievement %', 14),
            ('Weight', 10),
            ('Weighted Score', 15),
            ('# Employees', 12),
        ]
        
        for col_idx, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.cell_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # KPI display name mapping
        kpi_display_names = {
            'revenue': 'Revenue',
            'gross_profit': 'Gross Profit',
            'gross_profit_1': 'GP1',
            'num_new_customers': 'New Customers',
            'num_new_products': 'New Products',
            'num_new_combos': 'New Combos',
            'new_business_revenue': 'New Business Revenue',
        }
        
        # Currency KPIs (for formatting)
        currency_kpis = {'revenue', 'gross_profit', 'gross_profit_1', 'new_business_revenue'}
        
        # Data rows
        row = 5
        total_weighted_score = 0
        total_weight = 0
        
        for kpi in kpi_details:
            kpi_name = kpi.get('kpi_name', '')
            display_name = kpi_display_names.get(kpi_name, kpi_name.replace('_', ' ').title())
            actual = kpi.get('actual', 0)
            target_prorated = kpi.get('target_prorated', 0)
            target_annual = kpi.get('target_annual', 0)
            achievement = kpi.get('achievement', 0)
            weight = kpi.get('weight', 0)
            weighted_score = achievement * weight
            emp_count = kpi.get('employee_count', 0)
            
            is_currency = kpi_name in currency_kpis
            
            # KPI Type
            ws.cell(row=row, column=1, value=display_name)
            ws.cell(row=row, column=1).border = self.cell_border
            ws.cell(row=row, column=1).font = Font(bold=True)
            
            # Actual
            cell = ws.cell(row=row, column=2, value=actual)
            cell.border = self.cell_border
            cell.alignment = self.right_align
            if is_currency:
                cell.number_format = self.currency_format
            else:
                cell.number_format = '0.0'
            
            # Target (Prorated)
            cell = ws.cell(row=row, column=3, value=target_prorated)
            cell.border = self.cell_border
            cell.alignment = self.right_align
            if is_currency:
                cell.number_format = self.currency_format
            else:
                cell.number_format = '0.0'
            
            # Target (Annual)
            cell = ws.cell(row=row, column=4, value=target_annual)
            cell.border = self.cell_border
            cell.alignment = self.right_align
            if is_currency:
                cell.number_format = self.currency_format
            else:
                cell.number_format = '0.0'
            
            # Achievement %
            cell = ws.cell(row=row, column=5, value=achievement)
            cell.border = self.cell_border
            cell.alignment = self.right_align
            cell.number_format = '0.0'
            cell.fill = self._get_achievement_fill(achievement)
            
            # Weight
            cell = ws.cell(row=row, column=6, value=weight)
            cell.border = self.cell_border
            cell.alignment = self.center_align
            
            # Weighted Score
            cell = ws.cell(row=row, column=7, value=round(weighted_score, 1))
            cell.border = self.cell_border
            cell.alignment = self.right_align
            cell.number_format = '0.0'
            
            # # Employees
            cell = ws.cell(row=row, column=8, value=emp_count)
            cell.border = self.cell_border
            cell.alignment = self.center_align
            
            total_weighted_score += weighted_score
            total_weight += weight
            row += 1
        
        # Total row
        row += 1
        ws.cell(row=row, column=1, value="TOTAL / OVERALL")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).border = self.cell_border
        
        # Empty cells for columns 2-4
        for col in range(2, 5):
            ws.cell(row=row, column=col).border = self.cell_border
        
        # Overall Achievement
        overall_ach = overall_achievement_data.get('overall_achievement', 0)
        cell = ws.cell(row=row, column=5, value=overall_ach)
        cell.font = Font(bold=True)
        cell.border = self.cell_border
        cell.alignment = self.right_align
        cell.number_format = '0.0'
        cell.fill = self._get_achievement_fill(overall_ach)
        
        # Total Weight
        cell = ws.cell(row=row, column=6, value=total_weight)
        cell.font = Font(bold=True)
        cell.border = self.cell_border
        cell.alignment = self.center_align
        
        # Total Weighted Score
        cell = ws.cell(row=row, column=7, value=round(total_weighted_score, 1))
        cell.font = Font(bold=True)
        cell.border = self.cell_border
        cell.alignment = self.right_align
        cell.number_format = '0.0'
        
        # Empty cell for column 8
        ws.cell(row=row, column=8).border = self.cell_border
        
        # Verification note
        row += 2
        ws.cell(row=row, column=1, value=f"Verification: {total_weighted_score:.1f} / {total_weight} = {overall_ach:.1f}%")
        ws.cell(row=row, column=1).font = Font(italic=True, color='666666')
        
        # Freeze header
        ws.freeze_panes = 'A5'
    
    # =========================================================================
    # PIPELINE & FORECAST SHEET
    # =========================================================================
    
    def _create_pipeline_forecast_sheet(self, pipeline_metrics: Dict):
        """Create Pipeline & Forecast sheet with Revenue/GP/GP1 metrics."""
        ws = self.wb.create_sheet("Pipeline & Forecast")
        
        # Headers
        headers = [
            ('KPI Type', 15),
            ('Invoiced', 15),
            ('In-Period (KPI)', 18),
            ('Target', 15),
            ('Forecast', 15),
            ('GAP/Surplus', 15),
            ('Achievement %', 15),
            ('# Employees', 12),
        ]
        
        for col_idx, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.cell_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Data rows
        kpi_types = [
            ('Revenue', 'revenue'),
            ('Gross Profit', 'gross_profit'),
            ('GP1', 'gp1'),
        ]
        
        row = 2
        for kpi_name, kpi_key in kpi_types:
            data = pipeline_metrics.get(kpi_key, {})
            
            ws.cell(row=row, column=1, value=kpi_name)
            ws.cell(row=row, column=1).border = self.cell_border
            ws.cell(row=row, column=1).font = Font(bold=True)
            
            # Invoiced (from actuals)
            invoiced = data.get('invoiced', 0)
            cell = ws.cell(row=row, column=2, value=invoiced)
            cell.number_format = self.currency_format
            cell.border = self.cell_border
            cell.alignment = self.right_align
            
            # In-Period Backlog (KPI filtered)
            in_period = data.get('in_period_backlog', 0)
            cell = ws.cell(row=row, column=3, value=in_period)
            cell.number_format = self.currency_format
            cell.border = self.cell_border
            cell.alignment = self.right_align
            
            # Target
            target = data.get('target')
            cell = ws.cell(row=row, column=4, value=target if target else 'N/A')
            if target:
                cell.number_format = self.currency_format
            cell.border = self.cell_border
            cell.alignment = self.right_align
            
            # Forecast
            forecast = data.get('forecast')
            cell = ws.cell(row=row, column=5, value=forecast if forecast else 'N/A')
            if forecast:
                cell.number_format = self.currency_format
            cell.border = self.cell_border
            cell.alignment = self.right_align
            
            # GAP/Surplus
            gap = data.get('gap')
            cell = ws.cell(row=row, column=6, value=gap if gap is not None else 'N/A')
            if gap is not None:
                cell.number_format = self.currency_format
                # Color coding
                if gap >= 0:
                    cell.fill = self.achievement_good_fill
                else:
                    cell.fill = self.achievement_bad_fill
            cell.border = self.cell_border
            cell.alignment = self.right_align
            
            # Achievement %
            achievement = data.get('forecast_achievement')
            cell = ws.cell(row=row, column=7, value=f"{achievement:.1f}%" if achievement else 'N/A')
            cell.border = self.cell_border
            cell.alignment = self.right_align
            
            # # Employees
            emp_count = data.get('employee_count', 0)
            cell = ws.cell(row=row, column=8, value=emp_count)
            cell.border = self.cell_border
            cell.alignment = self.center_align
            
            row += 1
        
        # Summary row
        row += 1
        summary = pipeline_metrics.get('summary', {})
        ws.cell(row=row, column=1, value="Total Backlog (All Employees)")
        ws.cell(row=row, column=1).font = Font(bold=True, italic=True)
        
        ws.cell(row=row, column=2, value=f"Revenue: ${summary.get('total_backlog_revenue', 0):,.0f}")
        ws.cell(row=row, column=3, value=f"GP: ${summary.get('total_backlog_gp', 0):,.0f}")
        ws.cell(row=row, column=4, value=f"Orders: {int(summary.get('backlog_orders', 0)):,}")
        
        # Freeze header
        ws.freeze_panes = 'A2'
    
    # =========================================================================
    # SALES DETAIL SHEET
    # =========================================================================
    
    def _create_sales_detail_sheet(self, df: pd.DataFrame):
        """Create enhanced sales detail sheet with more columns."""
        if df.empty:
            return
        
        ws = self.wb.create_sheet("Sales Detail")
        
        # Define columns with width
        columns = [
            ('inv_date', 'Date', 12),
            ('inv_number', 'Invoice #', 18),
            ('oc_number', 'OC #', 18),
            ('customer_po_number', 'Customer PO', 15),
            ('customer', 'Customer', 30),
            ('product_pn', 'Product', 25),
            ('brand', 'Brand', 15),
            ('sales_by_split_usd', 'Revenue', 15),
            ('gross_profit_by_split_usd', 'GP', 15),
            ('gp1_by_split_usd', 'GP1', 15),
            ('split_rate_percent', 'Split %', 10),
            ('sales_name', 'Salesperson', 18),
        ]
        
        # Filter to available columns
        available_cols = [(c, h, w) for c, h, w in columns if c in df.columns]
        
        # Write headers
        for col_idx, (col_name, header, width) in enumerate(available_cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.cell_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Write data (limit to 10000 rows for performance)
        export_df = df[[c for c, _, _ in available_cols]].head(10000)
        
        for row_idx, row_data in enumerate(export_df.itertuples(index=False), 2):
            for col_idx, (col_name, _, _) in enumerate(available_cols, 1):
                value = row_data[col_idx - 1]
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.cell_border
                
                # Formatting
                if col_name in ['sales_by_split_usd', 'gross_profit_by_split_usd', 'gp1_by_split_usd']:
                    cell.number_format = self.currency_format
                    cell.alignment = self.right_align
                elif col_name == 'split_rate_percent':
                    cell.number_format = '0%'
                    cell.alignment = self.center_align
        
        # Add row count info
        if len(df) > 10000:
            row = len(export_df) + 3
            ws.cell(row=row, column=1, value=f"Note: Showing first 10,000 of {len(df):,} total rows")
            ws.cell(row=row, column=1).font = Font(italic=True, color='666666')
        
        # Freeze header
        ws.freeze_panes = 'A2'
    
    # =========================================================================
    # BACKLOG SUMMARY SHEET
    # =========================================================================
    
    def _create_backlog_summary_sheet(self, df: pd.DataFrame):
        """Create backlog summary by salesperson."""
        if df.empty:
            return
        
        ws = self.wb.create_sheet("Backlog Summary")
        
        columns = [
            ('sales_name', 'Salesperson', 25),
            ('total_backlog_revenue', 'Backlog Revenue', 18),
            ('total_backlog_gp', 'Backlog GP', 15),
            ('backlog_orders', '# Orders', 12),
            ('backlog_customers', '# Customers', 12),
        ]
        
        # Filter to available columns
        available_cols = [(c, h, w) for c, h, w in columns if c in df.columns]
        
        # Write headers
        for col_idx, (col_name, header, width) in enumerate(available_cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.cell_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Write data
        for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
            for col_idx, (col_name, _, _) in enumerate(available_cols, 1):
                try:
                    value = getattr(row_data, col_name) if hasattr(row_data, col_name) else df.iloc[row_idx - 2][col_name]
                except:
                    value = ''
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.cell_border
                
                if col_name in ['total_backlog_revenue', 'total_backlog_gp']:
                    cell.number_format = self.currency_format
                    cell.alignment = self.right_align
                elif col_name in ['backlog_orders', 'backlog_customers']:
                    cell.alignment = self.center_align
        
        # Add totals row
        total_row = len(df) + 2
        ws.cell(row=total_row, column=1, value="TOTAL")
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        ws.cell(row=total_row, column=1).border = self.cell_border
        
        for col_idx, (col_name, _, _) in enumerate(available_cols, 1):
            if col_name in ['total_backlog_revenue', 'total_backlog_gp', 'backlog_orders', 'backlog_customers']:
                total = df[col_name].sum() if col_name in df.columns else 0
                cell = ws.cell(row=total_row, column=col_idx, value=total)
                cell.font = Font(bold=True)
                cell.border = self.cell_border
                if col_name in ['total_backlog_revenue', 'total_backlog_gp']:
                    cell.number_format = self.currency_format
                    cell.alignment = self.right_align
                else:
                    cell.alignment = self.center_align
        
        # Freeze header
        ws.freeze_panes = 'A2'
    
    # =========================================================================
    # BACKLOG DETAIL SHEET
    # =========================================================================
    
    def _create_backlog_detail_sheet(self, df: pd.DataFrame, in_period_analysis: Dict = None):
        """Create backlog detail with all line items."""
        if df.empty:
            return
        
        ws = self.wb.create_sheet("Backlog Detail")
        
        columns = [
            ('oc_number', 'OC #', 18),
            ('oc_date', 'OC Date', 12),
            ('etd', 'ETD', 12),
            ('days_until_etd', 'Days to ETD', 12),
            ('customer', 'Customer', 30),
            ('customer_po_number', 'Customer PO', 15),
            ('product_pn', 'Product', 25),
            ('brand', 'Brand', 15),
            ('backlog_sales_by_split_usd', 'Amount', 15),
            ('backlog_gp_by_split_usd', 'GP', 15),
            ('split_rate_percent', 'Split %', 10),
            ('status', 'Status', 12),
            ('sales_name', 'Salesperson', 18),
        ]
        
        # Filter to available columns
        available_cols = [(c, h, w) for c, h, w in columns if c in df.columns]
        
        # Write headers
        for col_idx, (col_name, header, width) in enumerate(available_cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.cell_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Write data (limit to 5000 rows)
        export_df = df[[c for c, _, _ in available_cols]].head(5000)
        
        for row_idx, row_data in enumerate(export_df.itertuples(index=False), 2):
            for col_idx, (col_name, _, _) in enumerate(available_cols, 1):
                value = row_data[col_idx - 1]
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.cell_border
                
                # Formatting
                if col_name in ['backlog_sales_by_split_usd', 'backlog_gp_by_split_usd']:
                    cell.number_format = self.currency_format
                    cell.alignment = self.right_align
                elif col_name == 'split_rate_percent':
                    cell.number_format = '0%'
                    cell.alignment = self.center_align
                elif col_name == 'days_until_etd':
                    cell.alignment = self.center_align
                    # Color coding for overdue
                    if value is not None and isinstance(value, (int, float)) and value < 0:
                        cell.fill = self.achievement_bad_fill
        
        # Add summary info
        if in_period_analysis:
            row = len(export_df) + 3
            ws.cell(row=row, column=1, value="SUMMARY")
            ws.cell(row=row, column=1).font = Font(bold=True)
            
            row += 1
            ws.cell(row=row, column=1, value="Total Backlog:")
            ws.cell(row=row, column=2, value=f"${df['backlog_sales_by_split_usd'].sum():,.0f}")
            
            row += 1
            ws.cell(row=row, column=1, value="In-Period:")
            ws.cell(row=row, column=2, value=f"${in_period_analysis.get('total_value', 0):,.0f}")
            
            row += 1
            ws.cell(row=row, column=1, value="Overdue:")
            overdue_value = in_period_analysis.get('overdue_value', 0)
            ws.cell(row=row, column=2, value=f"${overdue_value:,.0f}")
            if overdue_value > 0:
                ws.cell(row=row, column=2).fill = self.achievement_bad_fill
        
        # Note if truncated
        if len(df) > 5000:
            row = len(export_df) + 8
            ws.cell(row=row, column=1, value=f"Note: Showing first 5,000 of {len(df):,} total rows")
            ws.cell(row=row, column=1).font = Font(italic=True, color='666666')
        
        # Freeze header
        ws.freeze_panes = 'A2'
    
    # =========================================================================
    # BACKLOG BY MONTH SHEET (LEGACY - Single Year)
    # =========================================================================
    
    def _create_backlog_by_month_sheet(self, df: pd.DataFrame):
        """Create backlog grouped by ETD month (single year)."""
        if df.empty:
            return
        
        ws = self.wb.create_sheet("Backlog by ETD")
        
        # Handle both 'month' (from prepare_backlog_by_month) and 'etd_month' (raw data)
        month_col = 'month' if 'month' in df.columns else 'etd_month'
        
        columns = [
            (month_col, 'Month', 12),
            ('etd_year', 'Year', 10),
            ('backlog_revenue', 'Backlog Revenue', 18),
            ('backlog_gp', 'Backlog GP', 15),
            ('order_count', '# Orders', 12),
        ]
        
        # Filter to available columns
        available_cols = [(c, h, w) for c, h, w in columns if c in df.columns]
        
        if not available_cols:
            return
        
        # Write headers
        for col_idx, (col_name, header, width) in enumerate(available_cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.cell_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Write data (filter out zero rows for cleaner output)
        export_df = df[df['backlog_revenue'] > 0] if 'backlog_revenue' in df.columns else df
        
        for row_idx, row_data in enumerate(export_df.itertuples(index=False), 2):
            for col_idx, (col_name, _, _) in enumerate(available_cols, 1):
                try:
                    value = export_df.iloc[row_idx - 2][col_name]
                except:
                    value = ''
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.cell_border
                
                if col_name in ['backlog_revenue', 'backlog_gp']:
                    cell.number_format = self.currency_format
                    cell.alignment = self.right_align
                elif col_name == 'order_count':
                    cell.alignment = self.center_align
        
        # Add totals row
        total_row = len(export_df) + 2
        ws.cell(row=total_row, column=1, value="TOTAL")
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        ws.cell(row=total_row, column=1).border = self.cell_border
        
        for col_idx, (col_name, _, _) in enumerate(available_cols, 1):
            if col_name in ['backlog_revenue', 'backlog_gp', 'order_count']:
                total = export_df[col_name].sum() if col_name in export_df.columns else 0
                cell = ws.cell(row=total_row, column=col_idx, value=total)
                cell.font = Font(bold=True)
                cell.border = self.cell_border
                if col_name in ['backlog_revenue', 'backlog_gp']:
                    cell.number_format = self.currency_format
                    cell.alignment = self.right_align
                else:
                    cell.alignment = self.center_align
        
        # Freeze header
        ws.freeze_panes = 'A2'
    
    # =========================================================================
    # BACKLOG BY MONTH SHEET - MULTI-YEAR (NEW v2.3.0)
    # =========================================================================
    
    def _create_backlog_by_month_sheet_multiyear(self, df: pd.DataFrame):
        """
        Create backlog grouped by ETD month with multi-year support.
        
        NEW v2.3.0: Shows all years in chronological order with year summary.
        """
        if df.empty:
            return
        
        ws = self.wb.create_sheet("Backlog by ETD")
        
        # Title
        ws.cell(row=1, column=1, value="Backlog by ETD Month (Multi-Year)")
        ws.cell(row=1, column=1).font = self.title_font
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        
        # Get unique years
        if 'etd_year' in df.columns:
            years = sorted(df['etd_year'].dropna().unique())
            ws.cell(row=2, column=1, value=f"Years: {', '.join(map(str, years))}")
            ws.cell(row=2, column=1).font = Font(italic=True, color='666666')
        
        # Headers
        columns = [
            ('year_month', 'Period', 12),
            ('etd_year', 'Year', 10),
            ('etd_month', 'Month', 10),
            ('backlog_revenue', 'Revenue', 18),
            ('backlog_gp', 'GP', 15),
            ('order_count', '# Orders', 12),
        ]
        
        # Filter to available columns
        available_cols = [(c, h, w) for c, h, w in columns if c in df.columns]
        
        if not available_cols:
            return
        
        header_row = 4
        for col_idx, (col_name, header, width) in enumerate(available_cols, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.cell_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Write data (sorted by sort_order if available)
        export_df = df.copy()
        if 'sort_order' in export_df.columns:
            export_df = export_df.sort_values('sort_order')
        
        # Filter out zero revenue rows
        if 'backlog_revenue' in export_df.columns:
            export_df = export_df[export_df['backlog_revenue'] > 0]
        
        row = header_row + 1
        current_year = None
        
        for _, row_data in export_df.iterrows():
            # Check for year change and add subtotal
            year = row_data.get('etd_year')
            if current_year is not None and year != current_year:
                # Add subtotal for previous year
                year_df = export_df[export_df['etd_year'] == current_year]
                if not year_df.empty:
                    ws.cell(row=row, column=1, value=f"Subtotal {int(current_year)}")
                    ws.cell(row=row, column=1).font = Font(bold=True, italic=True)
                    
                    for col_idx, (col_name, _, _) in enumerate(available_cols, 1):
                        if col_name in ['backlog_revenue', 'backlog_gp', 'order_count']:
                            subtotal = year_df[col_name].sum() if col_name in year_df.columns else 0
                            cell = ws.cell(row=row, column=col_idx, value=subtotal)
                            cell.font = Font(bold=True, italic=True)
                            if col_name in ['backlog_revenue', 'backlog_gp']:
                                cell.number_format = self.currency_format
                                cell.alignment = self.right_align
                    row += 1
            
            current_year = year
            
            # Write data row
            for col_idx, (col_name, _, _) in enumerate(available_cols, 1):
                value = row_data.get(col_name, '')
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = self.cell_border
                
                if col_name in ['backlog_revenue', 'backlog_gp']:
                    cell.number_format = self.currency_format
                    cell.alignment = self.right_align
                elif col_name == 'order_count':
                    cell.alignment = self.center_align
            
            row += 1
        
        # Final year subtotal
        if current_year is not None:
            year_df = export_df[export_df['etd_year'] == current_year]
            if not year_df.empty:
                ws.cell(row=row, column=1, value=f"Subtotal {int(current_year)}")
                ws.cell(row=row, column=1).font = Font(bold=True, italic=True)
                
                for col_idx, (col_name, _, _) in enumerate(available_cols, 1):
                    if col_name in ['backlog_revenue', 'backlog_gp', 'order_count']:
                        subtotal = year_df[col_name].sum() if col_name in year_df.columns else 0
                        cell = ws.cell(row=row, column=col_idx, value=subtotal)
                        cell.font = Font(bold=True, italic=True)
                        if col_name in ['backlog_revenue', 'backlog_gp']:
                            cell.number_format = self.currency_format
                            cell.alignment = self.right_align
                row += 1
        
        # Grand total
        row += 1
        ws.cell(row=row, column=1, value="GRAND TOTAL")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).border = self.cell_border
        
        for col_idx, (col_name, _, _) in enumerate(available_cols, 1):
            if col_name in ['backlog_revenue', 'backlog_gp', 'order_count']:
                total = export_df[col_name].sum() if col_name in export_df.columns else 0
                cell = ws.cell(row=row, column=col_idx, value=total)
                cell.font = Font(bold=True)
                cell.border = self.cell_border
                if col_name in ['backlog_revenue', 'backlog_gp']:
                    cell.number_format = self.currency_format
                    cell.alignment = self.right_align
                else:
                    cell.alignment = self.center_align
        
        # Freeze header
        ws.freeze_panes = f'A{header_row + 1}'
    
    # =========================================================================
    # COVER SHEET (LEGACY - kept for backward compatibility)
    # =========================================================================
    
    def _create_cover_sheet(
        self,
        metrics: Dict,
        filters: Dict,
        complex_kpis: Dict = None,
        yoy_metrics: Dict = None
    ):
        """Create cover page with KPI summary."""
        ws = self.wb.active
        ws.title = "Summary"
        
        row = 1
        
        # Title
        ws.cell(row=row, column=1, value="Salesperson Performance Report")
        ws.cell(row=row, column=1).font = self.title_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 2
        
        # Report Info
        ws.cell(row=row, column=1, value="Report Period:")
        ws.cell(row=row, column=2, value=f"{filters.get('period_type', 'YTD')} {filters.get('year', '')}")
        row += 1
        
        ws.cell(row=row, column=1, value="Date Range:")
        start_date = filters.get('start_date', '')
        end_date = filters.get('end_date', '')
        ws.cell(row=row, column=2, value=f"{start_date} to {end_date}")
        row += 1
        
        ws.cell(row=row, column=1, value="Generated:")
        ws.cell(row=row, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M'))
        row += 2
        
        # Main KPIs Section
        ws.cell(row=row, column=1, value="Key Performance Indicators")
        ws.cell(row=row, column=1).font = self.subtitle_font
        row += 1
        
        kpi_rows = [
            ("Total Revenue", f"${metrics.get('total_revenue', 0):,.0f}"),
            ("Gross Profit", f"${metrics.get('total_gp', 0):,.0f}"),
            ("GP1", f"${metrics.get('total_gp1', 0):,.0f}"),
            ("GP %", f"{metrics.get('gp_percent', 0):.1f}%"),
            ("Total Customers", f"{metrics.get('total_customers', 0):,}"),
            ("Total Invoices", f"{metrics.get('total_invoices', 0):,}"),
            ("Total Orders", f"{metrics.get('total_orders', 0):,}"),
        ]
        
        # Add achievement if available
        if metrics.get('revenue_achievement'):
            kpi_rows.append(("Revenue Achievement", f"{metrics['revenue_achievement']:.1f}%"))
        if metrics.get('gp_achievement'):
            kpi_rows.append(("GP Achievement", f"{metrics['gp_achievement']:.1f}%"))
        
        for label, value in kpi_rows:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=2).alignment = self.right_align
            row += 1
        
        row += 1
        
        # Complex KPIs Section (if provided)
        if complex_kpis:
            ws.cell(row=row, column=1, value="Complex KPIs")
            ws.cell(row=row, column=1).font = self.subtitle_font
            row += 1
            
            complex_rows = [
                ("New Customers", f"{complex_kpis.get('new_customer_count', 0):.1f}"),
                ("New Products", f"{complex_kpis.get('new_product_count', 0):.1f}"),
                ("New Business Revenue", f"${complex_kpis.get('new_business_revenue', 0):,.0f}"),
            ]
            
            for label, value in complex_rows:
                ws.cell(row=row, column=1, value=label)
                ws.cell(row=row, column=2, value=value)
                ws.cell(row=row, column=2).alignment = self.right_align
                row += 1
            
            row += 1
        
        # YoY Comparison (if provided)
        if yoy_metrics:
            ws.cell(row=row, column=1, value="Year-over-Year Comparison")
            ws.cell(row=row, column=1).font = self.subtitle_font
            row += 1
            
            yoy_rows = [
                ("Revenue YoY", yoy_metrics.get('total_revenue_yoy')),
                ("GP YoY", yoy_metrics.get('total_gp_yoy')),
                ("Customers YoY", yoy_metrics.get('total_customers_yoy')),
            ]
            
            for label, value in yoy_rows:
                ws.cell(row=row, column=1, value=label)
                if value is not None:
                    ws.cell(row=row, column=2, value=f"{value:+.1f}%")
                else:
                    ws.cell(row=row, column=2, value="N/A")
                ws.cell(row=row, column=2).alignment = self.right_align
                row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    # =========================================================================
    # SUMMARY SHEET (By Salesperson)
    # =========================================================================
    
    def _create_summary_sheet(self, df: pd.DataFrame):
        """Create salesperson summary sheet with formatting."""
        if df.empty:
            return
        
        ws = self.wb.create_sheet("By Salesperson")
        
        # Define columns to export (will filter to available)
        all_columns = [
            ('sales_name', 'Salesperson', 25),
            ('revenue', 'Revenue (USD)', 15),
            ('gross_profit', 'Gross Profit (USD)', 18),
            ('gp1', 'GP1 (USD)', 15),
            ('gp_percent', 'GP %', 10),
            ('customers', 'Customers', 12),
            ('invoices', 'Invoices', 10),
        ]
        
        # Add target columns if present
        if 'revenue_target' in df.columns:
            all_columns.append(('revenue_target', 'Revenue Target', 15))
        if 'gp_target' in df.columns:
            all_columns.append(('gp_target', 'GP Target', 15))
        
        # UPDATED v2.2.0: Prefer overall_achievement over revenue_achievement
        # overall_achievement = weighted average of all KPIs (same as Overview)
        if 'overall_achievement' in df.columns:
            all_columns.append(('overall_achievement', 'Achievement %', 14))
        elif 'revenue_achievement' in df.columns:
            all_columns.append(('revenue_achievement', 'Achievement %', 14))
        
        # Filter to available columns
        columns = [(c, h, w) for c, h, w in all_columns if c in df.columns]
        
        if not columns:
            return
        
        # Write headers
        for col_idx, (col_name, header, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.cell_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Write data
        for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
            for col_idx, (col_name, _, _) in enumerate(columns, 1):
                try:
                    value = df.iloc[row_idx - 2][col_name]
                except:
                    value = ''
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.cell_border
                
                # Apply formatting based on column type
                if col_name in ['revenue', 'gross_profit', 'gp1', 'revenue_target', 'gp_target']:
                    cell.number_format = self.currency_format
                    cell.alignment = self.right_align
                elif col_name in ['gp_percent', 'revenue_achievement', 'gp_achievement', 'overall_achievement']:
                    cell.alignment = self.right_align
                elif col_name in ['customers', 'invoices']:
                    cell.alignment = self.center_align
        
        # Add totals row
        total_row = len(df) + 2
        ws.cell(row=total_row, column=1, value="TOTAL")
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        ws.cell(row=total_row, column=1).border = self.cell_border
        
        for col_idx, (col_name, _, _) in enumerate(columns, 1):
            if col_name in ['revenue', 'gross_profit', 'gp1']:
                total = df[col_name].sum() if col_name in df.columns else 0
                cell = ws.cell(row=total_row, column=col_idx, value=total)
                cell.font = Font(bold=True)
                cell.border = self.cell_border
                cell.number_format = self.currency_format
                cell.alignment = self.right_align
            elif col_name in ['customers', 'invoices']:
                total = df[col_name].sum() if col_name in df.columns else 0
                cell = ws.cell(row=total_row, column=col_idx, value=total)
                cell.font = Font(bold=True)
                cell.border = self.cell_border
                cell.alignment = self.center_align
        
        # Add conditional formatting for achievement column
        # UPDATED v2.2.0: Support both overall_achievement and revenue_achievement
        achievement_col_name = None
        if 'overall_achievement' in df.columns:
            achievement_col_name = 'overall_achievement'
        elif 'revenue_achievement' in df.columns:
            achievement_col_name = 'revenue_achievement'
        
        if achievement_col_name:
            achievement_col = None
            for col_idx, (col_name, _, _) in enumerate(columns, 1):
                if col_name == achievement_col_name:
                    achievement_col = get_column_letter(col_idx)
                    break
            
            if achievement_col:
                # Red for <80%, Yellow for 80-100%, Green for >=100%
                ws.conditional_formatting.add(
                    f'{achievement_col}2:{achievement_col}{len(df) + 1}',
                    ColorScaleRule(
                        start_type='num', start_value=50, start_color='F8696B',
                        mid_type='num', mid_value=100, mid_color='FFEB84',
                        end_type='num', end_value=150, end_color='63BE7B'
                    )
                )
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    # =========================================================================
    # MONTHLY SHEET
    # =========================================================================
    
    def _create_monthly_sheet(self, df: pd.DataFrame):
        """Create monthly breakdown sheet."""
        if df.empty:
            return
        
        ws = self.wb.create_sheet("Monthly")
        
        columns = [
            ('invoice_month', 'Month', 12),
            ('revenue', 'Revenue (USD)', 15),
            ('gross_profit', 'Gross Profit (USD)', 18),
            ('gp1', 'GP1 (USD)', 15),
            ('gp_percent', 'GP %', 10),
            ('customer_count', 'Customers', 12),
            ('cumulative_revenue', 'Cumulative Revenue', 18),
            ('cumulative_gp', 'Cumulative GP', 15),
        ]
        
        # Write headers
        for col_idx, (col_name, header, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.cell_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Write data
        for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
            for col_idx, (col_name, _, _) in enumerate(columns, 1):
                try:
                    value = getattr(row_data, col_name) if hasattr(row_data, col_name) else df.iloc[row_idx - 2][col_name]
                except:
                    value = ''
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.cell_border
                
                if col_name in ['revenue', 'gross_profit', 'gp1', 'cumulative_revenue', 'cumulative_gp']:
                    cell.number_format = self.currency_format
                    cell.alignment = self.right_align
                elif col_name == 'gp_percent':
                    cell.alignment = self.right_align
                elif col_name in ['customer_count']:
                    cell.alignment = self.center_align
        
        # Freeze header
        ws.freeze_panes = 'A2'
    
    # =========================================================================
    # DETAIL SHEET (LEGACY)
    # =========================================================================
    
    def _create_detail_sheet(self, df: pd.DataFrame):
        """Create detailed transactions sheet."""
        if df.empty:
            return
        
        ws = self.wb.create_sheet("Details")
        
        # Select relevant columns
        detail_columns = [
            'inv_date', 'inv_number', 'sales_name', 'customer', 
            'product_pn', 'brand', 'sales_by_split_usd', 
            'gross_profit_by_split_usd', 'split_rate_percent'
        ]
        
        # Filter to available columns
        available_cols = [c for c in detail_columns if c in df.columns]
        
        if not available_cols:
            return
        
        export_df = df[available_cols].copy()
        
        # Rename columns for display
        rename_map = {
            'inv_date': 'Invoice Date',
            'inv_number': 'Invoice #',
            'sales_name': 'Salesperson',
            'customer': 'Customer',
            'product_pn': 'Product',
            'brand': 'Brand',
            'sales_by_split_usd': 'Revenue (USD)',
            'gross_profit_by_split_usd': 'GP (USD)',
            'split_rate_percent': 'Split %'
        }
        
        export_df.rename(columns=rename_map, inplace=True)
        
        # Write to sheet
        for col_idx, column in enumerate(export_df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.cell_border
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        
        for row_idx, row in enumerate(export_df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.cell_border
        
        # Freeze header
        ws.freeze_panes = 'A2'