# utils/safety_stock/export.py
"""
Export and reporting functions for Safety Stock Management
Version 2.2 - Updated to remove reorder_qty field
"""

import pandas as pd
import io
from datetime import datetime
from typing import Optional
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from sqlalchemy import text
from ..db import get_db_engine
from .permissions import get_user_role, log_action

logger = logging.getLogger(__name__)

# Excel formatting constants
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


def export_to_excel(
    df: pd.DataFrame,
    include_parameters: bool = True,
    include_metadata: bool = True
) -> io.BytesIO:
    """
    Export safety stock data to formatted Excel file
    
    Args:
        df: DataFrame to export
        include_parameters: Include calculation parameters sheet
        include_metadata: Include audit fields (created_by, etc.)
    
    Returns:
        BytesIO object containing Excel file
    """
    output = io.BytesIO()
    
    try:
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Main data columns - removed reorder_qty
            main_columns = [
                'pt_code', 'product_name', 'brand_name',
                'entity_code', 'entity_name',
                'customer_code', 'customer_name',
                'safety_stock_qty', 'reorder_point',
                'calculation_method', 'rule_type', 'status',
                'effective_from', 'effective_to',
                'priority_level', 'business_notes'
            ]
            
            # Add metadata columns if requested
            if include_metadata:
                main_columns.extend(['created_by', 'created_date', 'updated_by', 'updated_date'])
            
            # Filter to available columns
            export_columns = [col for col in main_columns if col in df.columns]
            main_df = df[export_columns].copy()
            
            # Format dates
            date_columns = ['effective_from', 'effective_to', 'created_date', 'updated_date']
            for col in date_columns:
                if col in main_df.columns:
                    main_df[col] = pd.to_datetime(main_df[col], errors='coerce').dt.strftime('%Y-%m-%d')
            
            # Fill NaN values for better display
            main_df['customer_code'] = main_df['customer_code'].fillna('ALL')
            main_df['customer_name'] = main_df['customer_name'].fillna('General Rule')
            
            # Write main sheet
            main_df.to_excel(writer, sheet_name='Safety Stock Levels', index=False)
            
            # Add parameters sheet if requested
            if include_parameters:
                param_df = _prepare_parameters_sheet(df)
                if not param_df.empty:
                    param_df.to_excel(writer, sheet_name='Calculation Parameters', index=False)
            
            # Format all sheets
            workbook = writer.book
            for sheet_name in workbook.sheetnames:
                _format_excel_sheet(workbook[sheet_name])
        
        output.seek(0)
        
        # Log export action
        log_action('EXPORT', f"Exported {len(df)} records to Excel")
        logger.info(f"Exported {len(df)} safety stock records by {get_user_role()}")
        
        return output
        
    except Exception as e:
        logger.error(f"Error exporting to Excel: {e}")
        raise


def _prepare_parameters_sheet(df: pd.DataFrame) -> pd.DataFrame:
    """Prepare calculation parameters sheet"""
    param_columns = [
        'pt_code', 'product_name', 'entity_code', 'customer_code',
        'calculation_method', 'lead_time_days', 'safety_days',
        'service_level_percent', 'avg_daily_demand', 'demand_std_deviation',
        'last_calculated_date'
    ]
    
    available_columns = [col for col in param_columns if col in df.columns]
    
    if not available_columns:
        return pd.DataFrame()
    
    param_df = df[available_columns].copy()
    
    # Only include rows with calculation parameters
    if 'calculation_method' in param_df.columns:
        param_df = param_df[param_df['calculation_method'].notna()]
        param_df = param_df[param_df['calculation_method'] != 'FIXED']
    
    # Format date
    if 'last_calculated_date' in param_df.columns:
        param_df['last_calculated_date'] = pd.to_datetime(
            param_df['last_calculated_date'], errors='coerce'
        ).dt.strftime('%Y-%m-%d %H:%M')
    
    return param_df


def _format_excel_sheet(worksheet, freeze_row: int = 2):
    """Apply formatting to Excel worksheet"""
    # Header formatting
    for cell in worksheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        # Set width with min/max limits
        adjusted_width = min(max(max_length + 2, 10), 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Add borders and alternate row colors
    for row_num, row in enumerate(worksheet.iter_rows(min_row=1), 1):
        for cell in row:
            cell.border = THIN_BORDER
            # Alternate row coloring (skip header)
            if row_num > 1 and row_num % 2 == 0:
                cell.fill = ALT_ROW_FILL
    
    # Freeze panes
    worksheet.freeze_panes = f'A{freeze_row}'


def create_upload_template(include_sample_data: bool = False) -> io.BytesIO:
    """
    Create Excel template for bulk upload
    
    Args:
        include_sample_data: Add sample rows
    
    Returns:
        BytesIO object containing template
    """
    output = io.BytesIO()
    
    try:
        # Template columns with descriptions - removed reorder_qty
        template_data = {
            'product_id': ['Required: Product ID from system'],
            'entity_id': ['Required: Entity/Company ID'],
            'customer_id': ['Optional: Customer ID (leave blank for general rule)'],
            'safety_stock_qty': ['Required: Safety Stock Quantity (>= 0)'],
            'reorder_point': ['Optional: Reorder trigger point'],
            'calculation_method': ['Optional: FIXED | DAYS_OF_SUPPLY | LEAD_TIME_BASED'],
            'lead_time_days': ['Optional: For LEAD_TIME_BASED method'],
            'safety_days': ['Optional: For DAYS_OF_SUPPLY method'],
            'service_level_percent': ['Optional: 90, 95, 98, 99 (for LEAD_TIME_BASED)'],
            'demand_std_deviation': ['Optional: For statistical calculation'],
            'avg_daily_demand': ['Optional: Average daily demand'],
            'effective_from': ['Required: Start date (YYYY-MM-DD)'],
            'effective_to': ['Optional: End date (YYYY-MM-DD)'],
            'priority_level': ['Optional: 1-9999 (default: 100)'],
            'business_notes': ['Optional: Notes/Comments']
        }
        
        df = pd.DataFrame(template_data)
        
        # Add sample data if requested
        if include_sample_data:
            sample_rows = [
                {
                    'product_id': 101,
                    'entity_id': 1,
                    'customer_id': '',
                    'safety_stock_qty': 100,
                    'reorder_point': 150,
                    'calculation_method': 'DAYS_OF_SUPPLY',
                    'lead_time_days': '',
                    'safety_days': 14,
                    'service_level_percent': '',
                    'demand_std_deviation': '',
                    'avg_daily_demand': 10,
                    'effective_from': datetime.now().strftime('%Y-%m-%d'),
                    'effective_to': '',
                    'priority_level': 100,
                    'business_notes': 'Example: Days of supply method'
                },
                {
                    'product_id': 102,
                    'entity_id': 1,
                    'customer_id': 5,
                    'safety_stock_qty': 75,
                    'reorder_point': 120,
                    'calculation_method': 'LEAD_TIME_BASED',
                    'lead_time_days': 7,
                    'safety_days': '',
                    'service_level_percent': 95,
                    'demand_std_deviation': 3.5,
                    'avg_daily_demand': 8,
                    'effective_from': datetime.now().strftime('%Y-%m-%d'),
                    'effective_to': '',
                    'priority_level': 50,
                    'business_notes': 'Example: Statistical method for customer'
                },
                {
                    'product_id': 103,
                    'entity_id': 2,
                    'customer_id': '',
                    'safety_stock_qty': 200,
                    'reorder_point': 250,
                    'calculation_method': 'FIXED',
                    'lead_time_days': '',
                    'safety_days': '',
                    'service_level_percent': '',
                    'demand_std_deviation': '',
                    'avg_daily_demand': '',
                    'effective_from': datetime.now().strftime('%Y-%m-%d'),
                    'effective_to': '',
                    'priority_level': 100,
                    'business_notes': 'Example: Manual fixed quantity'
                }
            ]
            
            sample_df = pd.DataFrame(sample_rows)
            df = pd.concat([df, sample_df], ignore_index=True)
        
        # Write to Excel
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Safety Stock Import', index=False)
            
            # Add instructions sheet
            instructions = _create_instructions()
            instructions_df = pd.DataFrame({'Instructions': instructions})
            instructions_df.to_excel(writer, sheet_name='Instructions', index=False)
            
            workbook = writer.book
            
            # Format main sheet
            main_sheet = workbook['Safety Stock Import']
            _format_excel_sheet(main_sheet)
            
            # Special formatting for description row
            for cell in main_sheet[1]:
                cell.font = Font(italic=True, color="FF0000", size=10)
                cell.alignment = Alignment(wrap_text=True)
            
            # Format instructions sheet
            inst_sheet = workbook['Instructions']
            inst_sheet.column_dimensions['A'].width = 100
            for row in inst_sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        output.seek(0)
        return output
        
    except Exception as e:
        logger.error(f"Error creating template: {e}")
        raise


def _create_instructions() -> list:
    """Create instructions for template"""
    return [
        'SAFETY STOCK BULK UPLOAD TEMPLATE - INSTRUCTIONS',
        '',
        '=== REQUIRED FIELDS ===',
        '• product_id: Product ID from the system',
        '• entity_id: Entity/Company ID',
        '• safety_stock_qty: Safety stock quantity (minimum 0)',
        '• effective_from: Start date in YYYY-MM-DD format',
        '',
        '=== CALCULATION METHODS ===',
        '',
        '1. FIXED',
        '   - Manual input, no calculation',
        '   - Use for: New products, special cases',
        '',
        '2. DAYS_OF_SUPPLY',
        '   - Formula: SS = safety_days × avg_daily_demand',
        '   - Required: safety_days',
        '   - Optional: avg_daily_demand (will calculate if not provided)',
        '',
        '3. LEAD_TIME_BASED',
        '   - Formula: SS = Z-score × √lead_time × std_deviation',
        '   - Required: lead_time_days, service_level_percent',
        '   - Service levels: 90, 95, 98, 99',
        '',
        '=== OPTIONAL FIELDS ===',
        '• reorder_point: Inventory level that triggers new purchase order',
        '• customer_id: Leave blank for general rules, or specify customer ID',
        '• effective_to: End date for the rule (blank = ongoing)',
        '• business_notes: Any additional context or notes',
        '',
        '=== PRIORITY RULES ===',
        '• Lower number = higher priority',
        '• General rules: 100 (default)',
        '• Customer-specific: 50 or lower',
        '',
        '=== IMPORTANT NOTES ===',
        '• Delete the first row (field descriptions) before uploading',
        '• Customer-specific rules override general rules',
        '• Date ranges cannot overlap for same product/entity/customer',
        '• Check the sample data rows for examples'
    ]


def generate_review_report(
    review_period_days: int = 30,
    entity_id: Optional[int] = None
) -> io.BytesIO:
    """
    Generate review report for safety stock
    
    Args:
        review_period_days: Period to analyze
        entity_id: Optional entity filter
    
    Returns:
        BytesIO object containing report
    """
    output = io.BytesIO()
    
    try:
        engine = get_db_engine()
        
        # Get report data
        summary_df = _get_report_summary(engine, review_period_days, entity_id)
        pending_df = _get_pending_reviews(engine, review_period_days, entity_id)
        recent_df = _get_recent_reviews(engine, review_period_days, entity_id)
        
        # Write to Excel
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            if not pending_df.empty:
                pending_df.to_excel(writer, sheet_name='Pending Reviews', index=False)
            
            if not recent_df.empty:
                recent_df.to_excel(writer, sheet_name='Recent Reviews', index=False)
            
            # Format sheets
            workbook = writer.book
            for sheet_name in workbook.sheetnames:
                _format_excel_sheet(workbook[sheet_name])
        
        output.seek(0)
        
        # Log action
        log_action('REPORT', f"Generated review report for {review_period_days} days")
        
        return output
        
    except Exception as e:
        logger.error(f"Error generating report: {e}")
        raise


def _get_report_summary(engine, days: int, entity_id: Optional[int]) -> pd.DataFrame:
    """Get summary statistics for report"""
    query = text("""
    SELECT 
        'Total Active Items' as Metric,
        COUNT(DISTINCT s.id) as Value
    FROM safety_stock_levels s
    WHERE s.delete_flag = 0 AND s.is_active = 1
    AND CURRENT_DATE() >= s.effective_from
    AND (s.effective_to IS NULL OR CURRENT_DATE() <= s.effective_to)
    """ + (" AND s.entity_id = :entity_id" if entity_id else "") + """
    
    UNION ALL
    
    SELECT 
        'Items Reviewed' as Metric,
        COUNT(DISTINCT ssr.safety_stock_level_id) as Value
    FROM safety_stock_reviews ssr
    WHERE ssr.review_date >= DATE_SUB(CURRENT_DATE(), INTERVAL :days DAY)
    """ + ("""
    AND ssr.safety_stock_level_id IN (
        SELECT id FROM safety_stock_levels WHERE entity_id = :entity_id
    )""" if entity_id else "") + """
    
    UNION ALL
    
    SELECT 
        'Pending Reviews' as Metric,
        COUNT(DISTINCT s.id) as Value
    FROM safety_stock_levels s
    LEFT JOIN safety_stock_parameters ssp ON s.id = ssp.safety_stock_level_id
    WHERE s.delete_flag = 0 AND s.is_active = 1
    AND (
        ssp.last_calculated_date IS NULL 
        OR ssp.last_calculated_date < DATE_SUB(CURRENT_DATE(), INTERVAL :days DAY)
    )
    """ + (" AND s.entity_id = :entity_id" if entity_id else ""))
    
    params = {'days': days}
    if entity_id:
        params['entity_id'] = entity_id
    
    with engine.connect() as conn:
        df = pd.read_sql(query, conn, params=params)
    
    return df


def _get_pending_reviews(engine, days: int, entity_id: Optional[int]) -> pd.DataFrame:
    """Get items pending review"""
    query = text("""
    SELECT 
        p.pt_code as 'Product Code',
        p.name as 'Product Name',
        e.company_code as 'Entity',
        s.safety_stock_qty as 'Current SS Qty',
        ssp.calculation_method as 'Method',
        CASE 
            WHEN ssp.last_calculated_date IS NULL THEN 'Never'
            ELSE DATEDIFF(CURRENT_DATE(), ssp.last_calculated_date)
        END as 'Days Since Calculation',
        s.priority_level as 'Priority'
    FROM safety_stock_levels s
    JOIN products p ON s.product_id = p.id
    JOIN companies e ON s.entity_id = e.id
    LEFT JOIN safety_stock_parameters ssp ON s.id = ssp.safety_stock_level_id
    LEFT JOIN (
        SELECT safety_stock_level_id, MAX(review_date) as last_review
        FROM safety_stock_reviews
        GROUP BY safety_stock_level_id
    ) ssr ON s.id = ssr.safety_stock_level_id
    WHERE s.delete_flag = 0 AND s.is_active = 1
    AND (
        ssr.last_review IS NULL 
        OR ssr.last_review < DATE_SUB(CURRENT_DATE(), INTERVAL :days DAY)
    )
    """ + (" AND s.entity_id = :entity_id" if entity_id else "") + """
    ORDER BY s.priority_level, p.pt_code
    LIMIT 100
    """)
    
    params = {'days': days}
    if entity_id:
        params['entity_id'] = entity_id
    
    with engine.connect() as conn:
        df = pd.read_sql(query, conn, params=params)
    
    return df


def _get_recent_reviews(engine, days: int, entity_id: Optional[int]) -> pd.DataFrame:
    """Get recent review history"""
    query = text("""
    SELECT 
        ssr.review_date as 'Review Date',
        p.pt_code as 'Product Code',
        ssr.old_safety_stock_qty as 'Old Qty',
        ssr.new_safety_stock_qty as 'New Qty',
        ssr.change_percentage as 'Change %',
        ssr.action_taken as 'Action',
        ssr.reviewed_by as 'Reviewed By',
        ssr.approved_by as 'Approved By'
    FROM safety_stock_reviews ssr
    JOIN safety_stock_levels s ON ssr.safety_stock_level_id = s.id
    JOIN products p ON s.product_id = p.id
    WHERE ssr.review_date >= DATE_SUB(CURRENT_DATE(), INTERVAL :days DAY)
    """ + (" AND s.entity_id = :entity_id" if entity_id else "") + """
    ORDER BY ssr.review_date DESC
    LIMIT 100
    """)
    
    params = {'days': days}
    if entity_id:
        params['entity_id'] = entity_id
    
    with engine.connect() as conn:
        df = pd.read_sql(query, conn, params=params)
    
    # Format date
    if not df.empty:
        df['Review Date'] = pd.to_datetime(df['Review Date']).dt.strftime('%Y-%m-%d')
    
    return df