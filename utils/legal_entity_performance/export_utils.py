# utils/legal_entity_performance/export_utils.py
"""
Export Utilities for Legal Entity Performance
Aligned with kpi_center_performance/export.py

VERSION: 2.0.0
- openpyxl formatted Excel export (synced with KPI center)
- CSV export for quick downloads
"""

import io
import logging
from datetime import datetime
from typing import Dict, Optional
import pandas as pd
import streamlit as st

from .constants import EXCEL_STYLES

logger = logging.getLogger(__name__)


class LegalEntityExport:
    """Handle CSV/Excel exports."""
    
    @staticmethod
    def to_csv(df: pd.DataFrame) -> bytes:
        return df.to_csv(index=False).encode('utf-8-sig')
    
    @staticmethod
    def to_excel(df: pd.DataFrame, sheet_name: str = 'Data') -> bytes:
        """Convert DataFrame to formatted Excel bytes using openpyxl."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        header_fill = PatternFill(
            start_color=EXCEL_STYLES['header_fill_color'],
            end_color=EXCEL_STYLES['header_fill_color'],
            fill_type='solid'
        )
        header_font = Font(bold=True, color=EXCEL_STYLES['header_font_color'], size=11)
        thin_border = Side(style='thin', color='000000')
        cell_border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
        
        # Write headers
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = cell_border
            ws.column_dimensions[get_column_letter(col_idx)].width = max(len(str(col_name)) + 4, 12)
        
        # Write data
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = cell_border
                
                col_name = df.columns[col_idx - 1].lower()
                if 'usd' in col_name or 'amount' in col_name or 'revenue' in col_name:
                    cell.number_format = EXCEL_STYLES['currency_format']
                    cell.alignment = Alignment(horizontal='right')
                elif 'percent' in col_name or col_name.endswith('_pct'):
                    cell.number_format = EXCEL_STYLES['percent_format']
        
        ws.freeze_panes = 'A2'
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    @staticmethod
    def render_download_button(df: pd.DataFrame, filename: str,
                                label: str = "📥 Export",
                                key: str = None):
        """Render download buttons for CSV and Excel."""
        if df.empty:
            st.info("No data to export")
            return
        
        col1, col2 = st.columns(2)
        
        with col1:
            csv_data = LegalEntityExport.to_csv(df)
            st.download_button(
                label=f"{label} CSV",
                data=csv_data,
                file_name=f"{filename}.csv",
                mime="text/csv",
                key=f"{key}_csv" if key else None
            )
        
        with col2:
            try:
                excel_data = LegalEntityExport.to_excel(df)
                st.download_button(
                    label=f"{label} Excel",
                    data=excel_data,
                    file_name=f"{filename}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"{key}_xlsx" if key else None
                )
            except ImportError:
                pass  # openpyxl not available
