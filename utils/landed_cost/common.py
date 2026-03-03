"""
Landed Cost - Common Utilities (v3.0)
Formatting, chart builders, heatmap helpers, constants, and Excel export.
Added: Landing charges analysis charts, cost decomposition, ratio heatmaps.
"""

import logging
from io import BytesIO
from typing import Any, Optional

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

logger = logging.getLogger(__name__)


# ================================================================
# Constants
# ================================================================

class LandedCostConstants:
    """Module-level constants."""
    MAX_MULTILINE_PRODUCTS = 15
    HEATMAP_TOP_N = 15
    TOP_PRODUCTS_N = 20
    SIGNIFICANT_CHANGE_PCT = 10.0
    LANDING_RATIO_ALERT_SHIFT = 5.0  # alert if ratio shifts >5 ppts

    COST_COLOR = "#4472C4"
    ACCENT_COLOR = "#ED7D31"
    ARRIVAL_COLOR = "#4472C4"
    OB_COLOR = "#ED7D31"
    TREND_COLORS = px.colors.qualitative.Set2

    # Cost breakdown colors
    PURCHASE_COLOR = "#4472C4"
    INTL_CHARGE_COLOR = "#ED7D31"
    LOCAL_CHARGE_COLOR = "#A5A5A5"
    IMPORT_TAX_COLOR = "#FFC000"
    LANDING_COLOR = "#E74C3C"


# ================================================================
# Formatting Helpers
# ================================================================

def format_usd(value: Any, decimals: int = 2) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "-"
    try:
        return f"${float(value):,.{decimals}f}"
    except (ValueError, TypeError):
        return str(value)


def format_usd4(value: Any) -> str:
    return format_usd(value, decimals=4)


def format_usd_smart(value: Any, max_decimals: int = 4, min_decimals: int = 2) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "-"
    try:
        num = float(value)
        formatted = f"{num:,.{max_decimals}f}"
        int_part, dec_part = formatted.split(".")
        dec_stripped = dec_part.rstrip("0")
        dec_final = dec_stripped.ljust(min_decimals, "0")
        return f"${int_part}.{dec_final}"
    except (ValueError, TypeError):
        return str(value)


def format_quantity(value: Any, decimals: int = 2) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "-"
    try:
        num = float(value)
        if num == int(num):
            return f"{int(num):,}"
        return f"{num:,.{decimals}f}".rstrip("0").rstrip(".")
    except (ValueError, TypeError):
        return str(value)


def format_pct_change(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "-"
    try:
        v = float(value)
        sign = "+" if v > 0 else ""
        return f"{sign}{v:.1f}%"
    except (ValueError, TypeError):
        return str(value)


def format_pct(value: Any, decimals: int = 1) -> str:
    """Format a percentage value (already in %, not as fraction)."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "-"
    try:
        return f"{float(value):.{decimals}f}%"
    except (ValueError, TypeError):
        return str(value)


def format_rate(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "-"
    try:
        return f"{float(value):,.4f}"
    except (ValueError, TypeError):
        return str(value)


def format_date(value: Any) -> str:
    if value is None:
        return "-"
    try:
        if isinstance(value, str):
            return value
        if hasattr(value, "strftime"):
            return value.strftime("%d/%m/%Y")
        return str(value)
    except Exception:
        return str(value)


def safe_get(data: dict, key: str, default: Any = None) -> Any:
    try:
        value = data.get(key, default)
        if pd.isna(value):
            return default
        return value
    except Exception:
        return default


# ================================================================
# Plotly Layout Defaults
# ================================================================

def _plotly_layout_defaults(fig, height: int = 400) -> go.Figure:
    fig.update_layout(
        height=height,
        margin=dict(l=20, r=20, t=40, b=20),
        font=dict(size=11),
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        xaxis=dict(showgrid=False),
        yaxis=dict(showgrid=True, gridcolor="rgba(0,0,0,0.06)"),
        hoverlabel=dict(bgcolor="white"),
    )
    return fig


# ================================================================
# Chart Builders — Existing (preserved)
# ================================================================

def build_cost_trend_chart(df: pd.DataFrame) -> Optional[go.Figure]:
    """Line chart: cost trend over years, multi-product or aggregated."""
    if df.empty or "cost_year" not in df.columns:
        return None

    products = df["pt_code"].nunique() if "pt_code" in df.columns else 0
    C = LandedCostConstants

    if 1 < products <= C.MAX_MULTILINE_PRODUCTS:
        fig = px.line(
            df.sort_values("cost_year"),
            x="cost_year",
            y="average_landed_cost_usd",
            color="pt_code",
            markers=True,
            labels={
                "cost_year": "Year",
                "average_landed_cost_usd": "Avg Cost (USD)",
                "pt_code": "Product",
            },
            color_discrete_sequence=C.TREND_COLORS,
        )
    else:
        agg = (
            df.groupby("cost_year")
            .agg(total_value=("total_landed_value_usd", "sum"),
                 total_qty=("total_quantity", "sum"))
            .reset_index()
        )
        agg["weighted_avg"] = agg["total_value"] / agg["total_qty"].replace(0, float("nan"))

        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=agg["cost_year"], y=agg["weighted_avg"],
            mode="lines+markers+text",
            text=[format_usd4(v) for v in agg["weighted_avg"]],
            textposition="top center",
            marker=dict(size=8, color=C.COST_COLOR),
            line=dict(width=2, color=C.COST_COLOR),
            name="Weighted Avg Cost",
            hovertemplate="Year: %{x}<br>Cost: $%{y:.4f}<extra></extra>",
        ))

    fig = _plotly_layout_defaults(fig, height=380)
    fig.update_xaxes(dtick=1, title_text="Year")
    fig.update_yaxes(title_text="USD")
    return fig


def build_cost_distribution_chart(
    df: pd.DataFrame, year: Optional[int] = None
) -> Optional[go.Figure]:
    """Box plot: cost distribution by brand."""
    data = df[df["cost_year"] == year] if year else df
    if data.empty or "brand" not in data.columns:
        return None

    top_brands = data.groupby("brand")["total_landed_value_usd"].sum().nlargest(12).index
    data = data[data["brand"].isin(top_brands)]

    fig = px.box(
        data, x="brand", y="average_landed_cost_usd",
        labels={"brand": "Brand", "average_landed_cost_usd": "Avg Cost (USD)"},
        color_discrete_sequence=[LandedCostConstants.COST_COLOR],
    )
    fig = _plotly_layout_defaults(fig, height=380)
    fig.update_xaxes(title_text="")
    return fig


def build_source_breakdown_chart(df: pd.DataFrame) -> Optional[go.Figure]:
    """Pie chart: Arrival vs Opening Balance quantities."""
    if df.empty:
        return None

    arr = df["arrival_quantity"].sum() if "arrival_quantity" in df.columns else 0
    ob = df["opening_balance_quantity"].sum() if "opening_balance_quantity" in df.columns else 0
    total = arr + ob
    if total == 0:
        return None

    fig = go.Figure(data=[go.Pie(
        labels=["Arrivals", "Opening Balance"],
        values=[arr, ob],
        marker=dict(colors=[LandedCostConstants.ARRIVAL_COLOR,
                            LandedCostConstants.OB_COLOR]),
        textinfo="label+percent",
        hovertemplate="%{label}: %{value:,.0f}<br>(%{percent})<extra></extra>",
        hole=0.4,
    )])
    fig = _plotly_layout_defaults(fig, height=380)
    fig.update_layout(showlegend=True)
    return fig


def build_yoy_comparison_table(yoy_df: pd.DataFrame) -> pd.DataFrame:
    """Build comparison table between 2 most recent years.
    Uses INNER join on (product_id, entity_id) for consistent product matching.
    """
    if yoy_df.empty or yoy_df["cost_year"].nunique() < 2:
        return pd.DataFrame()

    years = sorted(yoy_df["cost_year"].unique(), reverse=True)
    curr, prev = years[0], years[1]

    join_keys = ["product_id", "entity_id"]

    curr_data = yoy_df[yoy_df["cost_year"] == curr][
        join_keys + ["pt_code", "product_pn", "brand", "legal_entity",
         "average_landed_cost_usd", "total_quantity", "total_landed_value_usd"]
    ].copy()
    prev_data = yoy_df[yoy_df["cost_year"] == prev][
        join_keys + ["average_landed_cost_usd", "total_quantity", "total_landed_value_usd"]
    ].copy()

    curr_data.rename(columns={
        "average_landed_cost_usd": f"cost_{curr}",
        "total_quantity": f"qty_{curr}",
        "total_landed_value_usd": f"value_{curr}",
    }, inplace=True)
    prev_data.rename(columns={
        "average_landed_cost_usd": f"cost_{prev}",
        "total_quantity": f"qty_{prev}",
        "total_landed_value_usd": f"value_{prev}",
    }, inplace=True)

    merged = curr_data.merge(prev_data, on=join_keys, how="inner")
    merged["yoy_change_usd"] = merged[f"cost_{curr}"] - merged[f"cost_{prev}"]
    merged["yoy_change_pct"] = (
        merged["yoy_change_usd"] / merged[f"cost_{prev}"].replace(0, float("nan")) * 100
    )

    return merged.sort_values("yoy_change_pct", ascending=False, na_position="last")

def build_yoy_breakdown_table(
    yoy_df: pd.DataFrame, breakdown_df: pd.DataFrame
) -> pd.DataFrame:
    """Build YoY comparison table with purchase/landing decomposition."""
    if yoy_df.empty or yoy_df["cost_year"].nunique() < 2:
        return pd.DataFrame()

    years = sorted(yoy_df["cost_year"].unique(), reverse=True)
    curr, prev = years[0], years[1]

    # Base comparison from main view
    base = build_yoy_comparison_table(yoy_df)
    if base.empty or breakdown_df.empty:
        return base

    # Merge breakdown data for both years
    for yr in [curr, prev]:
        yr_bd = breakdown_df[breakdown_df["cost_year"] == yr][
            ["product_id", "entity_id",
             "avg_purchase_cost_usd", "avg_landing_charge_usd", "landing_ratio_pct"]
        ].copy()
        yr_bd.rename(columns={
            "avg_purchase_cost_usd": f"purchase_{yr}",
            "avg_landing_charge_usd": f"landing_{yr}",
            "landing_ratio_pct": f"ratio_{yr}",
        }, inplace=True)
        base = base.merge(yr_bd, on=["product_id", "entity_id"], how="left")

    # Compute YoY changes for purchase and landing
    base[f"purchase_yoy_pct"] = (
        (base[f"purchase_{curr}"] - base[f"purchase_{prev}"])
        / base[f"purchase_{prev}"].replace(0, float("nan")) * 100
    ).round(1)
    base[f"landing_yoy_pct"] = (
        (base[f"landing_{curr}"] - base[f"landing_{prev}"])
        / base[f"landing_{prev}"].replace(0, float("nan")) * 100
    ).round(1)
    base["ratio_shift"] = (
        base[f"ratio_{curr}"].fillna(0) - base[f"ratio_{prev}"].fillna(0)
    ).round(1)

    return base


# ================================================================
# Chart Builders — NEW: Landing Charges Analysis
# ================================================================

def build_cost_composition_chart(bd_df: pd.DataFrame) -> Optional[go.Figure]:
    """Stacked bar: Purchase Cost / Intl / Local / Tax by year."""
    if bd_df.empty or "cost_year" not in bd_df.columns:
        return None

    C = LandedCostConstants
    agg = bd_df.groupby("cost_year").agg(
        purchase=("total_purchase_value_usd", "sum"),
        intl=("total_international_charge_usd", "sum"),
        local=("total_local_charge_usd", "sum"),
        tax=("total_import_tax_usd", "sum"),
    ).reset_index().sort_values("cost_year")

    fig = go.Figure()
    for col, name, color in [
        ("purchase", "Purchase Cost", C.PURCHASE_COLOR),
        ("intl", "International Charges", C.INTL_CHARGE_COLOR),
        ("local", "Local Charges", C.LOCAL_CHARGE_COLOR),
        ("tax", "Import Tax", C.IMPORT_TAX_COLOR),
    ]:
        fig.add_trace(go.Bar(
            x=agg["cost_year"].astype(str), y=agg[col],
            name=name, marker_color=color,
            hovertemplate=f"{name}: $%{{y:,.0f}}<extra></extra>",
        ))

    fig.update_layout(barmode="stack", legend=dict(orientation="h", yanchor="bottom", y=1.02))
    fig = _plotly_layout_defaults(fig, height=400)
    fig.update_xaxes(title_text="Year")
    fig.update_yaxes(title_text="USD")
    return fig


def build_landing_ratio_trend_chart(bd_df: pd.DataFrame) -> Optional[go.Figure]:
    """Line chart: landing ratio % trend over years."""
    if bd_df.empty or bd_df["cost_year"].nunique() < 2:
        return None

    C = LandedCostConstants
    agg = bd_df.groupby("cost_year").agg(
        total_purchase=("total_purchase_value_usd", "sum"),
        total_landed=("total_landed_value_usd", "sum"),
    ).reset_index().sort_values("cost_year")

    agg["landing_ratio"] = (
        (agg["total_landed"] - agg["total_purchase"])
        / agg["total_purchase"].replace(0, float("nan")) * 100
    ).round(2)

    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=agg["cost_year"], y=agg["landing_ratio"],
        mode="lines+markers+text",
        text=[f"{v:.1f}%" for v in agg["landing_ratio"]],
        textposition="top center",
        marker=dict(size=8, color=C.LANDING_COLOR),
        line=dict(width=2, color=C.LANDING_COLOR),
        name="Landing Ratio",
        hovertemplate="Year: %{x}<br>Ratio: %{y:.1f}%<extra></extra>",
    ))

    fig = _plotly_layout_defaults(fig, height=320)
    fig.update_xaxes(dtick=1, title_text="Year")
    fig.update_yaxes(title_text="Landing Ratio %")
    return fig


def build_landing_by_ship_method_chart(sm_df: pd.DataFrame) -> Optional[go.Figure]:
    """Grouped bar: landing charge components by ship method."""
    if sm_df.empty:
        return None

    C = LandedCostConstants
    sm_df = sm_df.sort_values("total_landing_charges_usd", ascending=False).head(8)

    fig = go.Figure()
    for col, name, color in [
        ("total_international_charge_usd", "International", C.INTL_CHARGE_COLOR),
        ("total_local_charge_usd", "Local", C.LOCAL_CHARGE_COLOR),
        ("total_import_tax_usd", "Import Tax", C.IMPORT_TAX_COLOR),
    ]:
        fig.add_trace(go.Bar(
            x=sm_df["ship_method"], y=sm_df[col],
            name=name, marker_color=color,
            hovertemplate=f"{name}: $%{{y:,.0f}}<extra></extra>",
        ))

    fig.update_layout(barmode="group", legend=dict(orientation="h", yanchor="bottom", y=1.02))
    fig = _plotly_layout_defaults(fig, height=380)
    fig.update_xaxes(title_text="")
    fig.update_yaxes(title_text="USD")
    return fig


def build_landing_by_country_chart(country_df: pd.DataFrame) -> Optional[go.Figure]:
    """Horizontal bar: landing ratio by vendor country."""
    if country_df.empty:
        return None

    C = LandedCostConstants
    top = country_df.dropna(subset=["landing_ratio_pct"]).nlargest(12, "total_landed_value_usd")
    if top.empty:
        return None

    top = top.sort_values("landing_ratio_pct", ascending=True)

    fig = go.Figure()
    fig.add_trace(go.Bar(
        y=top["vendor_country"], x=top["landing_ratio_pct"],
        orientation="h",
        marker=dict(
            color=top["landing_ratio_pct"],
            colorscale="YlOrRd",
            showscale=True,
            colorbar=dict(title="%"),
        ),
        text=[f"{v:.1f}%" for v in top["landing_ratio_pct"]],
        textposition="outside",
        hovertemplate="Country: %{y}<br>Landing Ratio: %{x:.1f}%<br>"
                      "<extra></extra>",
    ))

    h = max(300, len(top) * 32 + 80)
    fig = _plotly_layout_defaults(fig, height=h)
    fig.update_xaxes(title_text="Landing Ratio %")
    fig.update_yaxes(title_text="")
    return fig


# ================================================================
# Chart Builders — NEW: Detail Dialog Decomposition
# ================================================================

def build_cost_decomposition_bar(breakdown: dict) -> Optional[go.Figure]:
    """Horizontal stacked bar: Purchase Cost vs Landing Charges (for dialog)."""
    purchase = breakdown.get("total_purchase_value_usd") or 0
    landing = breakdown.get("total_landing_charges_usd") or 0
    total = purchase + landing
    if total <= 0:
        return None

    C = LandedCostConstants
    pct_p = purchase / total * 100
    pct_l = landing / total * 100

    fig = go.Figure()
    fig.add_trace(go.Bar(
        y=["Cost"], x=[purchase], name="Purchase Cost",
        orientation="h", marker_color=C.PURCHASE_COLOR,
        text=[f"Purchase {pct_p:.0f}%"], textposition="inside",
        hovertemplate=f"Purchase Cost: ${purchase:,.2f} ({pct_p:.1f}%)<extra></extra>",
    ))
    fig.add_trace(go.Bar(
        y=["Cost"], x=[landing], name="Landing Charges",
        orientation="h", marker_color=C.LANDING_COLOR,
        text=[f"Landing {pct_l:.0f}%"], textposition="inside",
        hovertemplate=f"Landing Charges: ${landing:,.2f} ({pct_l:.1f}%)<extra></extra>",
    ))

    fig.update_layout(
        barmode="stack", showlegend=True,
        legend=dict(orientation="h", yanchor="bottom", y=1.02),
        height=100, margin=dict(l=20, r=20, t=30, b=10),
        font=dict(size=11),
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        xaxis=dict(showgrid=False, showticklabels=False),
        yaxis=dict(showgrid=False, showticklabels=False),
    )
    return fig


def build_landing_donut_chart(breakdown: dict) -> Optional[go.Figure]:
    """Donut chart: International / Local / Import Tax breakdown."""
    C = LandedCostConstants
    intl = breakdown.get("total_international_charge_usd") or 0
    local = breakdown.get("total_local_charge_usd") or 0
    tax = breakdown.get("total_import_tax_usd") or 0
    total = intl + local + tax
    if total <= 0:
        return None

    fig = go.Figure(data=[go.Pie(
        labels=["International", "Local", "Import Tax"],
        values=[intl, local, tax],
        marker=dict(colors=[C.INTL_CHARGE_COLOR, C.LOCAL_CHARGE_COLOR, C.IMPORT_TAX_COLOR]),
        textinfo="label+percent",
        hovertemplate="%{label}: $%{value:,.2f}<br>(%{percent})<extra></extra>",
        hole=0.45,
    )])
    fig = _plotly_layout_defaults(fig, height=280)
    fig.update_layout(showlegend=True,
                      legend=dict(orientation="h", yanchor="bottom", y=-0.15))
    return fig


# ================================================================
# Heatmap Builders — Existing (preserved)
# ================================================================

def build_brand_year_heatmap(df: pd.DataFrame) -> Optional[go.Figure]:
    """Heatmap: Brand × Year weighted average cost."""
    if df.empty or df["brand"].nunique() < 2 or df["cost_year"].nunique() < 2:
        return None

    top = df.groupby("brand")["total_landed_value_usd"].sum().nlargest(
        LandedCostConstants.HEATMAP_TOP_N).index
    heat_df = df[df["brand"].isin(top)].copy()

    pivot = heat_df.pivot_table(
        values="total_landed_value_usd", index="brand",
        columns="cost_year", aggfunc="sum", fill_value=0,
    )
    qty_pivot = heat_df.pivot_table(
        values="total_quantity", index="brand",
        columns="cost_year", aggfunc="sum", fill_value=0,
    )
    avg_pivot = (pivot / qty_pivot.replace(0, float("nan"))).round(4)
    avg_pivot = avg_pivot.sort_index()

    fig = go.Figure(data=go.Heatmap(
        z=avg_pivot.values,
        x=[str(c) for c in avg_pivot.columns],
        y=avg_pivot.index.tolist(),
        colorscale="YlOrRd",
        text=[[format_usd4(v) if pd.notna(v) else "" for v in row]
              for row in avg_pivot.values],
        texttemplate="%{text}",
        textfont=dict(size=10),
        hovertemplate="Brand: %{y}<br>Year: %{x}<br>Avg Cost: $%{z:.4f}<extra></extra>",
        colorbar=dict(title="USD"),
    ))
    h = max(400, len(avg_pivot) * 32 + 80)
    fig = _plotly_layout_defaults(fig, height=h)
    fig.update_layout(xaxis_title="Cost Year", yaxis_title="", showlegend=False)
    return fig


def build_entity_year_heatmap(df: pd.DataFrame) -> Optional[go.Figure]:
    """Heatmap: Entity × Year total value."""
    if df.empty or df["legal_entity"].nunique() < 2 or df["cost_year"].nunique() < 2:
        return None

    pivot = df.pivot_table(
        values="total_landed_value_usd", index="legal_entity",
        columns="cost_year", aggfunc="sum", fill_value=0,
    )
    pivot = pivot.sort_index()

    fig = go.Figure(data=go.Heatmap(
        z=pivot.values,
        x=[str(c) for c in pivot.columns],
        y=pivot.index.tolist(),
        colorscale="Blues",
        text=[[format_usd(v) if v > 0 else "" for v in row] for row in pivot.values],
        texttemplate="%{text}",
        textfont=dict(size=10),
        hovertemplate="Entity: %{y}<br>Year: %{x}<br>Value: $%{z:,.2f}<extra></extra>",
        colorbar=dict(title="USD"),
    ))
    h = max(300, len(pivot) * 40 + 80)
    fig = _plotly_layout_defaults(fig, height=h)
    fig.update_layout(xaxis_title="Cost Year", yaxis_title="", showlegend=False)
    return fig


# ================================================================
# Heatmap Builders — NEW: Landing Ratio
# ================================================================

def build_landing_ratio_heatmap(
    bd_df: pd.DataFrame, index_col: str = "brand", title_label: str = "Brand"
) -> Optional[go.Figure]:
    """Heatmap: brand (or entity) × Year landing ratio %."""
    if bd_df.empty or bd_df[index_col].nunique() < 2 or bd_df["cost_year"].nunique() < 2:
        return None

    C = LandedCostConstants

    # Aggregate purchase and landed values
    agg = bd_df.groupby([index_col, "cost_year"]).agg(
        purchase=("total_purchase_value_usd", "sum"),
        landed=("total_landed_value_usd", "sum"),
    ).reset_index()
    agg["ratio"] = ((agg["landed"] - agg["purchase"]) / agg["purchase"].replace(0, float("nan")) * 100).round(1)

    # If too many items, take top N by value
    if agg[index_col].nunique() > C.HEATMAP_TOP_N:
        top_items = agg.groupby(index_col)["landed"].sum().nlargest(C.HEATMAP_TOP_N).index
        agg = agg[agg[index_col].isin(top_items)]

    pivot = agg.pivot_table(values="ratio", index=index_col, columns="cost_year", fill_value=None)
    pivot = pivot.sort_index()

    fig = go.Figure(data=go.Heatmap(
        z=pivot.values,
        x=[str(c) for c in pivot.columns],
        y=pivot.index.tolist(),
        colorscale="RdYlGn_r",
        text=[[f"{v:.1f}%" if pd.notna(v) else "" for v in row] for row in pivot.values],
        texttemplate="%{text}",
        textfont=dict(size=10),
        hovertemplate=f"{title_label}: %{{y}}<br>Year: %{{x}}<br>Landing Ratio: %{{z:.1f}}%<extra></extra>",
        colorbar=dict(title="%"),
    ))
    h = max(350, len(pivot) * 32 + 80)
    fig = _plotly_layout_defaults(fig, height=h)
    fig.update_layout(xaxis_title="Cost Year", yaxis_title="", showlegend=False)
    return fig


# ================================================================
# Excel Export
# ================================================================

def create_excel_download(df: pd.DataFrame) -> bytes:
    """Create formatted Excel file from export DataFrame."""
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Landed Cost", startrow=2)
        ws = writer.sheets["Landed Cost"]
        ncols = len(df.columns)

        # Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        title_cell = ws.cell(row=1, column=1, value="LANDED COST REPORT")
        title_cell.font = Font(bold=True, size=13)
        title_cell.alignment = Alignment(horizontal="center")

        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        for col_idx in range(1, ncols + 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Auto width
        for idx, col in enumerate(df.columns):
            max_len = max(
                df[col].astype(str).map(len).max() if len(df) > 0 else 0,
                len(str(col)),
            ) + 2
            ws.column_dimensions[get_column_letter(idx + 1)].width = min(max_len, 45)

    return output.getvalue()