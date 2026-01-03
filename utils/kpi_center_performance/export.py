# utils/kpi_center_performance/export.py
"""
Formatted Excel Export for KPI Center Performance

"""

import logging
from datetime import datetime
from io import BytesIO
from typing import Dict, List, Optional
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill, NamedStyle
)
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule

from .constants import EXCEL_STYLES

logger = logging.getLogger(__name__)


class KPICenterExport:
    """
    Excel report generator for KPI Center performance.
    
    Usage:
        exporter = KPICenterExport()
        excel_bytes = exporter.create_comprehensive_report(
            metrics=metrics,
            complex_kpis=complex_kpis,
            ...
        )
        
        st.download_button(
            label="Download Report",
            data=excel_bytes,
            file_name="kpi_center_performance.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    """
    
    def __init__(self):
        """Initialize with default styles."""
        self.wb = None
        self._init_styles()
    
    def _init_styles(self):
        """Initialize reusable styles."""
        self.header_fill = PatternFill(
            start_color=EXCEL_STYLES['header_fill_color'],
            end_color=EXCEL_STYLES['header_fill_color'],
            fill_type='solid'
        )
        
        self.header_font = Font(
            bold=True,
            color=EXCEL_STYLES['header_font_color'],
            size=11
        )
        
        self.title_font = Font(bold=True, size=16)
        self.subtitle_font = Font(bold=True, size=12)
        self.normal_font = Font(size=11)
        
        thin_border = Side(style='thin', color='000000')
        self.cell_border = Border(
            left=thin_border,
            right=thin_border,
            top=thin_border,
            bottom=thin_border
        )
        
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.right_align = Alignment(horizontal='right', vertical='center')
        self.left_align = Alignment(horizontal='left', vertical='center')
        
        self.currency_format = EXCEL_STYLES['currency_format']
        self.percent_format = EXCEL_STYLES['percent_format']
    
    # =========================================================================
    # COMPREHENSIVE REPORT
    # =========================================================================
    
    def create_comprehensive_report(
        self,
        # Summary metrics
        metrics: Dict,
        complex_kpis: Dict,
        pipeline_metrics: Dict,
        filters: Dict,
        yoy_metrics: Dict = None,
        
        # KPI Center & Monthly data
        kpi_center_summary_df: pd.DataFrame = None,
        monthly_df: pd.DataFrame = None,
        
        # Sales Detail
        sales_detail_df: pd.DataFrame = None,
        
        # Backlog data
        backlog_summary_df: pd.DataFrame = None,
        backlog_detail_df: pd.DataFrame = None,
        backlog_by_month_df: pd.DataFrame = None,
    ) -> BytesIO:
        """
        Create comprehensive Excel report with all performance data.
        
        Sheets:
        1. Summary - KPIs, filters, period info
        2. Pipeline & Forecast - Revenue/GP/GP1 forecast vs target
        3. By KPI Center - Summary by KPI Center
        4. Monthly Trend - Monthly breakdown
        5. Sales Detail - Transaction list
        6. Backlog Summary - Backlog by KPI Center
        7. Backlog Detail - Backlog line items
        8. Backlog by ETD - Backlog grouped by ETD month
        """
        self.wb = Workbook()
        
        # Create sheets
        self._create_summary_sheet(metrics, complex_kpis, filters, yoy_metrics)
        self._create_pipeline_sheet(pipeline_metrics, filters)
        
        if kpi_center_summary_df is not None and not kpi_center_summary_df.empty:
            self._create_kpi_center_summary_sheet(kpi_center_summary_df)
        
        if monthly_df is not None and not monthly_df.empty:
            self._create_monthly_sheet(monthly_df)
        
        if sales_detail_df is not None and not sales_detail_df.empty:
            self._create_sales_detail_sheet(sales_detail_df)
        
        if backlog_summary_df is not None and not backlog_summary_df.empty:
            self._create_backlog_summary_sheet(backlog_summary_df)
        
        if backlog_detail_df is not None and not backlog_detail_df.empty:
            self._create_backlog_detail_sheet(backlog_detail_df)
        
        if backlog_by_month_df is not None and not backlog_by_month_df.empty:
            self._create_backlog_by_month_sheet(backlog_by_month_df)
        
        # Remove default empty sheet if exists
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']
        
        # Save to BytesIO
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        
        logger.info("KPI Center Excel report created successfully")
        return output
    
    # =========================================================================
    # SUMMARY SHEET
    # =========================================================================
    
    def _create_summary_sheet(
        self,
        metrics: Dict,
        complex_kpis: Dict,
        filters: Dict,
        yoy_metrics: Dict = None
    ):
        """Create summary sheet with KPIs."""
        ws = self.wb.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = "KPI Center Performance Report"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:D1')
        
        # Generation info
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A2'].font = Font(italic=True, size=10)
        
        # Filters section
        row = 4
        ws[f'A{row}'] = "FILTERS"
        ws[f'A{row}'].font = self.subtitle_font
        row += 1
        
        ws[f'A{row}'] = "Period:"
        ws[f'B{row}'] = f"{filters.get('period_type', 'YTD')} {filters.get('year', '')}"
        row += 1
        
        ws[f'A{row}'] = "Date Range:"
        start = filters.get('start_date', '')
        end = filters.get('end_date', '')
        ws[f'B{row}'] = f"{start} to {end}"
        row += 1
        
        ws[f'A{row}'] = "KPI Centers:"
        ws[f'B{row}'] = len(filters.get('kpi_center_ids', []))
        row += 2
        
        # Performance metrics
        ws[f'A{row}'] = "PERFORMANCE METRICS"
        ws[f'A{row}'].font = self.subtitle_font
        row += 1
        
        perf_data = [
            ("Revenue", f"${metrics.get('total_revenue', 0):,.0f}"),
            ("Gross Profit", f"${metrics.get('total_gp', 0):,.0f}"),
            ("GP1", f"${metrics.get('total_gp1', 0):,.0f}"),
            ("GP %", f"{metrics.get('gp_percent', 0):.1f}%"),
            ("GP1 %", f"{metrics.get('gp1_percent', 0):.1f}%"),
            ("Customers", f"{metrics.get('total_customers', 0):,}"),
            ("Orders", f"{metrics.get('total_orders', 0):,}"),
        ]
        
        for label, value in perf_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # YoY comparison
        if yoy_metrics:
            row += 1
            ws[f'A{row}'] = "YEAR-OVER-YEAR"
            ws[f'A{row}'].font = self.subtitle_font
            row += 1
            
            yoy_data = [
                ("Revenue YoY", yoy_metrics.get('total_revenue_yoy')),
                ("GP YoY", yoy_metrics.get('total_gp_yoy')),
                ("Customers YoY", yoy_metrics.get('total_customers_yoy')),
            ]
            
            for label, value in yoy_data:
                ws[f'A{row}'] = label
                if value is not None:
                    ws[f'B{row}'] = f"{value:+.1f}%"
                else:
                    ws[f'B{row}'] = "N/A"
                row += 1
        
        # New Business metrics
        if complex_kpis:
            row += 1
            ws[f'A{row}'] = "NEW BUSINESS"
            ws[f'A{row}'].font = self.subtitle_font
            row += 1
            
            new_biz_data = [
                ("New Customers", complex_kpis.get('num_new_customers', 0)),
                ("New Products", complex_kpis.get('num_new_products', 0)),
                ("New Business Revenue", f"${complex_kpis.get('new_business_revenue', 0):,.0f}"),
            ]
            
            for label, value in new_biz_data:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value if isinstance(value, str) else f"{value:.0f}"
                row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    # =========================================================================
    # PIPELINE SHEET
    # =========================================================================
    
    def _create_pipeline_sheet(self, pipeline_metrics: Dict, filters: Dict):
        """Create Pipeline & Forecast sheet."""
        ws = self.wb.create_sheet("Pipeline & Forecast")
        
        ws['A1'] = "Pipeline & Forecast"
        ws['A1'].font = self.title_font
        
        row = 3
        
        # Headers
        headers = ['Metric', 'Invoiced', 'In-Period Backlog', 'Target', 'Forecast', 'GAP', 'Achievement %']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.cell_border
            cell.alignment = self.center_align
        
        row += 1
        
        # Data rows
        for metric_name, key in [('Revenue', 'revenue'), ('Gross Profit', 'gross_profit'), ('GP1', 'gp1')]:
            data = pipeline_metrics.get(key, {})
            
            ws.cell(row=row, column=1, value=metric_name).border = self.cell_border
            
            for col, field in enumerate(['invoiced', 'in_period_backlog', 'target', 'forecast', 'gap'], 2):
                value = data.get(field, 0)
                cell = ws.cell(row=row, column=col, value=value if value else 0)
                cell.number_format = self.currency_format
                cell.border = self.cell_border
                cell.alignment = self.right_align
            
            # Achievement %
            achievement = data.get('forecast_achievement')
            cell = ws.cell(row=row, column=7, value=achievement if achievement else 0)
            cell.number_format = '0.0%' if achievement and achievement < 10 else '0.0'
            cell.border = self.cell_border
            cell.alignment = self.right_align
            
            row += 1
        
        # Adjust widths
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    # =========================================================================
    # KPI CENTER SUMMARY SHEET
    # =========================================================================
    
    def _create_kpi_center_summary_sheet(self, df: pd.DataFrame):
        """Create KPI Center summary sheet."""
        if df.empty:
            return
        
        ws = self.wb.create_sheet("By KPI Center")
        
        columns = [
            ('kpi_center', 'KPI Center', 25),
            ('kpi_type', 'Type', 12),
            ('revenue', 'Revenue (USD)', 15),
            ('gross_profit', 'Gross Profit (USD)', 18),
            ('gp1', 'GP1 (USD)', 15),
            ('gp_percent', 'GP %', 10),
            ('customers', 'Customers', 12),
            ('invoices', 'Invoices', 10),
        ]
        
        # Add target columns if present
        if 'revenue_target' in df.columns:
            columns.append(('revenue_target', 'Revenue Target', 15))
            columns.append(('revenue_achievement', 'Achievement %', 14))
        
        # Filter to available columns
        columns = [(c, h, w) for c, h, w in columns if c in df.columns]
        
        if not columns:
            return
        
        # Write headers
        for col_idx, (col_name, header, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.cell_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Write data
        for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
            for col_idx, (col_name, _, _) in enumerate(columns, 1):
                try:
                    value = df.iloc[row_idx - 2][col_name]
                except:
                    value = ''
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.cell_border
                
                if col_name in ['revenue', 'gross_profit', 'gp1', 'revenue_target']:
                    cell.number_format = self.currency_format
                    cell.alignment = self.right_align
                elif col_name in ['gp_percent', 'revenue_achievement']:
                    cell.alignment = self.right_align
                elif col_name in ['customers', 'invoices']:
                    cell.alignment = self.center_align
        
        # Add totals row
        total_row = len(df) + 2
        ws.cell(row=total_row, column=1, value="TOTAL")
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        ws.cell(row=total_row, column=1).border = self.cell_border
        
        for col_idx, (col_name, _, _) in enumerate(columns, 1):
            if col_name in ['revenue', 'gross_profit', 'gp1']:
                total = df[col_name].sum() if col_name in df.columns else 0
                cell = ws.cell(row=total_row, column=col_idx, value=total)
                cell.font = Font(bold=True)
                cell.border = self.cell_border
                cell.number_format = self.currency_format
                cell.alignment = self.right_align
        
        ws.freeze_panes = 'A2'
    
    # =========================================================================
    # MONTHLY SHEET
    # =========================================================================
    
    def _create_monthly_sheet(self, df: pd.DataFrame):
        """Create monthly breakdown sheet."""
        if df.empty:
            return
        
        ws = self.wb.create_sheet("Monthly")
        
        columns = [
            ('invoice_month', 'Month', 12),
            ('revenue', 'Revenue (USD)', 15),
            ('gross_profit', 'Gross Profit (USD)', 18),
            ('gp1', 'GP1 (USD)', 15),
            ('gp_percent', 'GP %', 10),
            ('customer_count', 'Customers', 12),
            ('cumulative_revenue', 'Cumulative Revenue', 18),
            ('cumulative_gp', 'Cumulative GP', 15),
        ]
        
        # Write headers
        for col_idx, (col_name, header, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.cell_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Write data
        for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
            for col_idx, (col_name, _, _) in enumerate(columns, 1):
                try:
                    value = getattr(row_data, col_name) if hasattr(row_data, col_name) else df.iloc[row_idx - 2][col_name]
                except:
                    value = ''
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.cell_border
                
                if col_name in ['revenue', 'gross_profit', 'gp1', 'cumulative_revenue', 'cumulative_gp']:
                    cell.number_format = self.currency_format
                    cell.alignment = self.right_align
                elif col_name == 'gp_percent':
                    cell.alignment = self.right_align
        
        ws.freeze_panes = 'A2'
    
    # =========================================================================
    # SALES DETAIL SHEET
    # =========================================================================
    
    def _create_sales_detail_sheet(self, df: pd.DataFrame):
        """Create sales detail sheet."""
        if df.empty:
            return
        
        ws = self.wb.create_sheet("Sales Detail")
        
        # Select relevant columns
        detail_columns = [
            'inv_date', 'inv_number', 'kpi_center', 'customer', 
            'product_pn', 'brand', 'sales_by_kpi_center_usd', 
            'gross_profit_by_kpi_center_usd', 'split_rate_percent'
        ]
        
        available_cols = [c for c in detail_columns if c in df.columns]
        
        if not available_cols:
            return
        
        export_df = df[available_cols].copy()
        
        rename_map = {
            'inv_date': 'Invoice Date',
            'inv_number': 'Invoice #',
            'kpi_center': 'KPI Center',
            'customer': 'Customer',
            'product_pn': 'Product',
            'brand': 'Brand',
            'sales_by_kpi_center_usd': 'Revenue (USD)',
            'gross_profit_by_kpi_center_usd': 'GP (USD)',
            'split_rate_percent': 'Split %'
        }
        
        export_df.rename(columns=rename_map, inplace=True)
        
        # Write to sheet
        for col_idx, column in enumerate(export_df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.cell_border
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        
        for row_idx, row in enumerate(export_df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.cell_border
        
        ws.freeze_panes = 'A2'
    
    # =========================================================================
    # BACKLOG SHEETS
    # =========================================================================
    
    def _create_backlog_summary_sheet(self, df: pd.DataFrame):
        """Create backlog summary by KPI Center sheet."""
        if df.empty:
            return
        
        ws = self.wb.create_sheet("Backlog Summary")
        
        columns = [
            ('kpi_center', 'KPI Center', 25),
            ('total_backlog_usd', 'Backlog (USD)', 15),
            ('total_backlog_gp_usd', 'Backlog GP (USD)', 15),
            ('backlog_orders', 'Orders', 10),
            ('backlog_customers', 'Customers', 12),
        ]
        
        columns = [(c, h, w) for c, h, w in columns if c in df.columns]
        
        for col_idx, (col_name, header, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.cell_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        for row_idx, _ in enumerate(df.itertuples(index=False), 2):
            for col_idx, (col_name, _, _) in enumerate(columns, 1):
                try:
                    value = df.iloc[row_idx - 2][col_name]
                except:
                    value = ''
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.cell_border
                
                if 'usd' in col_name:
                    cell.number_format = self.currency_format
                    cell.alignment = self.right_align
        
        ws.freeze_panes = 'A2'
    
    def _create_backlog_detail_sheet(self, df: pd.DataFrame):
        """Create backlog detail sheet."""
        if df.empty:
            return
        
        ws = self.wb.create_sheet("Backlog Detail")
        
        columns = [
            ('oc_number', 'OC #', 12),
            ('etd', 'ETD', 12),
            ('customer', 'Customer', 25),
            ('product_pn', 'Product', 20),
            ('kpi_center', 'KPI Center', 20),
            ('backlog_by_kpi_center_usd', 'Amount (USD)', 15),
            ('backlog_gp_by_kpi_center_usd', 'GP (USD)', 12),
            ('days_until_etd', 'Days to ETD', 12),
            ('pending_type', 'Status', 15),
        ]
        
        columns = [(c, h, w) for c, h, w in columns if c in df.columns]
        
        for col_idx, (col_name, header, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.cell_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Limit to 5000 rows for performance
        export_df = df.head(5000)
        
        for row_idx, _ in enumerate(export_df.itertuples(index=False), 2):
            for col_idx, (col_name, _, _) in enumerate(columns, 1):
                try:
                    value = export_df.iloc[row_idx - 2][col_name]
                except:
                    value = ''
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.cell_border
                
                if 'usd' in col_name:
                    cell.number_format = self.currency_format
                    cell.alignment = self.right_align
        
        ws.freeze_panes = 'A2'
    
    def _create_backlog_by_month_sheet(self, df: pd.DataFrame):
        """Create backlog by ETD month sheet."""
        if df.empty:
            return
        
        ws = self.wb.create_sheet("Backlog by ETD")
        
        columns = [
            ('etd_year', 'Year', 10),
            ('etd_month', 'Month', 10),
            ('backlog_orders', 'Orders', 10),
            ('backlog_usd', 'Backlog (USD)', 15),
            ('backlog_gp_usd', 'Backlog GP (USD)', 15),
        ]
        
        columns = [(c, h, w) for c, h, w in columns if c in df.columns]
        
        for col_idx, (col_name, header, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.cell_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        for row_idx, _ in enumerate(df.itertuples(index=False), 2):
            for col_idx, (col_name, _, _) in enumerate(columns, 1):
                try:
                    value = df.iloc[row_idx - 2][col_name]
                except:
                    value = ''
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.cell_border
                
                if 'usd' in col_name:
                    cell.number_format = self.currency_format
                    cell.alignment = self.right_align
        
        ws.freeze_panes = 'A2'